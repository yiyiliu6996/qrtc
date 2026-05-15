"""
╔══════════════════════════════════════════════════════╗
║              QrTC  V2  —  by Claude                  ║
║  Tạo QR chuyển khoản hàng loạt qua Telegram Bot     ║
║  Hỗ trợ: 1→nhiều, nhiều→1, whitelist, báo cáo Excel ║
╚══════════════════════════════════════════════════════╝
"""

import asyncio
import json
import logging
import os
import re
import sqlite3
import tempfile
import time
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional
from urllib.parse import urlencode

import aiohttp
from aiogram import Bot, Dispatcher, F
from aiogram.client.default import DefaultBotProperties
from aiogram.enums import ParseMode
from aiogram.filters import Command
from aiogram.types import FSInputFile, Message, CallbackQuery
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from PIL import Image

# ─────────────────────────── ĐƯỜNG DẪN FILE ──────────────────────────────────

# Trên Railway dùng /app/data để persist qua redeploy
# Local dùng thư mục hiện tại
_DATA_DIR = os.environ.get("RAILWAY_VOLUME_MOUNT_PATH", ".")
DB_PATH          = os.path.join(_DATA_DIR, "outqrbot.db")
CANCEL_LOG_PATH  = os.path.join(_DATA_DIR, "cancel_log.jsonl")
CONFIG_PATH      = "config.json"  # chỉ dùng local

# ─────────────────────────── LOGGING ─────────────────────────────────────────

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
logger = logging.getLogger("outqrbot")

# ─────────────────────────── ĐỌC CONFIG TỪ config.json ───────────────────────

def _load_config() -> dict:
    """
    Load config theo thứ tự ưu tiên:
    1. Environment variables (Railway / production)
    2. config.json (local development)
    """
    # Thử đọc từ environment variables trước (Railway)
    if os.environ.get("BOT_TOKEN"):
        ids_raw = os.environ.get("SUPER_ADMIN_IDS", "")
        super_ids = [int(x.strip()) for x in ids_raw.split(",") if x.strip().isdigit()]
        return {
            "BOT_TOKEN":                os.environ["BOT_TOKEN"].strip(),
            "SUPER_ADMIN_IDS":          super_ids,
            "NOTIFY_GROUP_ID":          int(os.environ.get("NOTIFY_GROUP_ID", "0") or 0),
            "COLLECT_GROUP_ID":         int(os.environ.get("COLLECT_GROUP_ID", "0") or 0),
            "DEFAULT_TRANSFER_CONTENT": os.environ.get("DEFAULT_TRANSFER_CONTENT", ""),
            "FORM_COOLDOWN_SECONDS":    int(os.environ.get("FORM_COOLDOWN_SECONDS", "3") or 3),
            "BANK_WHITELIST_ENABLED":   os.environ.get("BANK_WHITELIST_ENABLED", "false").lower() == "true",
            "WHITELIST_MODE":           os.environ.get("WHITELIST_MODE", "blacklist"),
        }

    # Fallback: đọc config.json (local)
    if not os.path.exists(CONFIG_PATH):
        print("\n❌  Chưa có file config.json!")
        print("   Chạy lệnh sau để thiết lập lần đầu:\n")
        print("       python setup.py\n")
        raise SystemExit(1)
    with open(CONFIG_PATH, "r", encoding="utf-8") as f:
        cfg = json.load(f)
    return cfg

_cfg = _load_config()

BOT_TOKEN:               str      = _cfg["BOT_TOKEN"]
SUPER_ADMIN_IDS:         set[int] = set(_cfg["SUPER_ADMIN_IDS"])
NOTIFY_GROUP_ID:         int      = _cfg["NOTIFY_GROUP_ID"]
COLLECT_GROUP_ID:        int      = _cfg.get("COLLECT_GROUP_ID", 0)
BANK_WHITELIST_ENABLED:  bool     = bool(_cfg.get("BANK_WHITELIST_ENABLED", False))
WHITELIST_MODE:          str      = _cfg.get("WHITELIST_MODE", "blacklist")  # "blacklist" hoặc "strict"
DEFAULT_TRANSFER_CONTENT: str     = _cfg.get("DEFAULT_TRANSFER_CONTENT", "")
FORM_COOLDOWN_SECONDS:   int      = _cfg.get("FORM_COOLDOWN_SECONDS", 3)

# Validate
if not BOT_TOKEN:
    raise RuntimeError("BOT_TOKEN trong config.json đang trống.")
if not SUPER_ADMIN_IDS:
    raise RuntimeError("SUPER_ADMIN_IDS trong config.json đang trống.")

# Internal
_last_form_time: dict[int, float] = defaultdict(float)
_cooldown_lock = asyncio.Lock()
_callback_locks: dict[str, asyncio.Lock] = {}   # order_code → Lock
_callback_locks_mu = asyncio.Lock()              # protect dict access

async def _get_order_lock(order_code: str) -> asyncio.Lock:
    async with _callback_locks_mu:
        if order_code not in _callback_locks:
            _callback_locks[order_code] = asyncio.Lock()
        return _callback_locks[order_code]

# ─────────────────────────── DB HELPERS ──────────────────────────────────────

def now_local() -> datetime:
    """Trả về giờ hiện tại theo GMT+7 (Asia/Ho_Chi_Minh)."""
    try:
        from zoneinfo import ZoneInfo
        return datetime.now(ZoneInfo("Asia/Ho_Chi_Minh")).replace(tzinfo=None)
    except Exception:
        # Fallback: UTC+7 thủ công
        from datetime import timezone, timedelta
        tz_vn = timezone(timedelta(hours=7))
        return datetime.now(tz_vn).replace(tzinfo=None)


def db_connect() -> sqlite3.Connection:
    conn = sqlite3.connect(DB_PATH, timeout=10, check_same_thread=False)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA foreign_keys=ON")
    return conn


def init_db() -> None:
    conn = db_connect()
    cur = conn.cursor()

    cur.execute("""
    CREATE TABLE IF NOT EXISTS activated_chats (
        chat_id      INTEGER PRIMARY KEY,
        activated_by INTEGER NOT NULL,
        activated_at TEXT NOT NULL
    )""")

    cur.execute("""
    CREATE TABLE IF NOT EXISTS orders (
        order_code          TEXT PRIMARY KEY,
        created_at          TEXT NOT NULL,
        chat_id             INTEGER NOT NULL,
        group_name          TEXT,
        creator_user_id     INTEGER NOT NULL,
        creator_name        TEXT,
        creator_message_id  INTEGER,
        qr_message_id       INTEGER,
        collect_message_id  INTEGER,
        completed_at        TEXT,
        completed_by_name   TEXT,
        sender_bank         TEXT,
        sender_account      TEXT,
        sender_name         TEXT,
        total_amount        INTEGER NOT NULL,
        total_qr            INTEGER NOT NULL,
        status              TEXT NOT NULL DEFAULT 'active'
    )""")
    # Migration: thêm cột mới nếu chưa có
    for col in [
        "ALTER TABLE orders ADD COLUMN creator_message_id INTEGER",
        "ALTER TABLE orders ADD COLUMN qr_message_id INTEGER",
        "ALTER TABLE orders ADD COLUMN button_message_id INTEGER",
        "ALTER TABLE orders ADD COLUMN collect_message_id INTEGER",
        "ALTER TABLE orders ADD COLUMN completed_at TEXT",
        "ALTER TABLE orders ADD COLUMN completed_by_name TEXT",
        "ALTER TABLE orders ADD COLUMN creator_username TEXT",
    ]:
        try:
            cur.execute(col)
        except Exception:
            pass

    cur.execute("""
    CREATE TABLE IF NOT EXISTS receivers (
        id               INTEGER PRIMARY KEY AUTOINCREMENT,
        order_code       TEXT NOT NULL,
        receiver_index   INTEGER NOT NULL,
        receiver_bank    TEXT,
        receiver_account TEXT,
        receiver_name    TEXT,
        amount           INTEGER NOT NULL,
        actual_amount    INTEGER,
        content          TEXT,
        message_id       INTEGER,
        FOREIGN KEY(order_code) REFERENCES orders(order_code)
    )""")
    try:
        cur.execute("ALTER TABLE receivers ADD COLUMN actual_amount INTEGER")
    except Exception:
        pass

    cur.execute("""
    CREATE TABLE IF NOT EXISTS order_messages (
        chat_id     INTEGER NOT NULL,
        message_id  INTEGER NOT NULL,
        order_code  TEXT NOT NULL,
        PRIMARY KEY(chat_id, message_id)
    )""")

    cur.execute("""
    CREATE TABLE IF NOT EXISTS bot_admins (
        chat_id   INTEGER NOT NULL,
        user_id   INTEGER NOT NULL,
        added_by  INTEGER,
        added_at  TEXT,
        PRIMARY KEY(chat_id, user_id)
    )""")

    # Whitelist tài khoản người nhận
    cur.execute("""
    CREATE TABLE IF NOT EXISTS receiver_whitelist (
        id          INTEGER PRIMARY KEY AUTOINCREMENT,
        bank        TEXT,
        device_code TEXT,
        account     TEXT NOT NULL UNIQUE,
        name        TEXT,
        is_active   INTEGER NOT NULL DEFAULT 1,
        added_at    TEXT NOT NULL,
        added_by    INTEGER
    )""")
    try:
        cur.execute("ALTER TABLE receiver_whitelist ADD COLUMN device_code TEXT")
    except Exception:
        pass

    # Global admin (thay thế bot_admins per-group)
    cur.execute("""
    CREATE TABLE IF NOT EXISTS global_admins (
        user_id    INTEGER PRIMARY KEY,
        added_by   INTEGER,
        added_at   TEXT NOT NULL,
        note       TEXT
    )""")

    # Migration: copy bot_admins → global_admins (bỏ trùng)
    cur.execute("""
        INSERT OR IGNORE INTO global_admins (user_id, added_by, added_at)
        SELECT DISTINCT user_id, added_by, added_at
        FROM bot_admins
    """)

    # Thêm cột pending_update để track ai đang chờ file txt
    try:
        cur.execute("ALTER TABLE orders ADD COLUMN creator_username TEXT")
    except Exception:
        pass

    conn.commit()
    conn.close()
    logger.info("DB khởi tạo xong: %s", DB_PATH)

# ─────────────────────────── BANK DATA ───────────────────────────────────────

# ── Bảng mã ngân hàng: mọi cách gõ → mã chuẩn VietQR ────────────────────────
BANK_ALIAS: dict[str, str] = {
    # Vietcombank
    "vcb":"vcb","vietcombank":"vcb","vietcom":"vcb","ngoaithuong":"vcb",
    # Vietinbank
    "vietinbank":"icb","icb":"icb","ctg":"icb","congthuong":"icb","viettinbank":"icb",
    # BIDV
    "bidv":"bidv","dautuphattrien":"bidv",
    # Agribank
    "agribank":"agribank","agri":"agribank","vbard":"agribank","nongnghiep":"agribank",
    # MB Bank
    "mb":"mb","mbbank":"mb","quandoi":"mb","militarybank":"mb",
    # MBV — Ngân hàng TNHH MTV Việt Nam Hiện Đại (BIN 970414)
    "mbv":"mbv","vietthienhai":"mbv","hiendai":"mbv",
    "nganhanghiendai":"mbv","vietnamhiendai":"mbv",
    # Techcombank
    "tcb":"tcb","techcombank":"tcb","techcom":"tcb","kythuong":"tcb",
    # ACB
    "acb":"acb","achau":"acb","asicommercial":"acb",
    # VPBank
    "vpbank":"vpb","vpb":"vpb","thinhvuong":"vpb",
    # TPBank
    "tpbank":"tpb","tpb":"tpb","tienphong":"tpb",
    # SHB
    "shb":"shb","saigonhanoi":"shb","saigonh":"shb",
    # Sacombank
    "sacombank":"stb","stb":"stb","sacom":"stb","saigonthongtin":"stb",
    # VIB
    "vib":"vib","quocte":"vib","internationalvn":"vib",
    # HDBank
    "hdbank":"hdb","hdb":"hdb","phattrientp":"hdb",
    # MSB
    "msb":"msb","maritimebank":"msb","hanghaibank":"msb","hanghai":"msb",
    # LPBank
    "lpbank":"lpb","lpb":"lpb","locphat":"lpb","lienphat":"lpb",
    # OCB
    "ocb":"ocb","phuongdong":"ocb","orientcommercial":"ocb",
    # SeABank
    "seabank":"seab","seab":"seab","dongnamai":"seab","dongnam":"seab",
    # Eximbank
    "eximbank":"eib","eib":"eib","exim":"eib","xuatnhapkhau":"eib",
    # Bac A Bank
    "bacabank":"bab","bab":"bab","baca":"bab","baca":"bab",
    # NCB
    "ncb":"ncb","quocdan":"ncb","nationalcitizen":"ncb","nncb":"ncb",
    # SCB
    "scb":"scb","saigonbank":"scb",
    # ABBank
    "abbank":"abb","abb":"abb","anbinh":"abb",
    # VietABank
    "vietabank":"vab","vab":"vab","vieta":"vab","vietnama":"vab",
    # NamABank
    "namabank":"nab","nab":"nab","nama":"nab","namabank":"nab",
    # PVComBank
    "pvcombank":"pvcb","pvcb":"pvcb","pvcom":"pvcb","daukhitoanc":"pvcb",
    # KienLongBank / Umee
    "kienlongbank":"klb","klb":"klb","kienlong":"klb","umee":"klb",
    # VietBank
    "vietbank":"vietbank","thuongtin":"vietbank",
    # BaoViet Bank
    "baovietbank":"bvb","bvb":"bvb","baoviet":"bvb",
    # CBBank
    "cbbank":"cbb","cbb":"cbb","xaydung":"cbb",
    # OceanBank
    "oceanbank":"oceanbank","ocean":"oceanbank","daididuong":"oceanbank",
    # GPBank
    "gpbank":"gpb","gpb":"gpb","daukhi":"gpb",
    # Shinhan Bank
    "shinhanbank":"shbvn","shbvn":"shbvn","shinhan":"shbvn","shinhanvn":"shbvn",
    # Woori
    "woori":"woori","wooribank":"woori",
    # HSBC
    "hsbc":"hsbc",
    # KBank
    "kbank":"kbank","kasikorn":"kbank",
    # CIMB
    "cimb":"cimb","cimbbank":"cimb",
    # PublicBank
    "publicbank":"pbvn","pbvn":"pbvn","publicvn":"pbvn",
    # HongLeong
    "hongleong":"hlbvn","hlbvn":"hlbvn",
    # Standard Chartered
    "standardchartered":"scvn","scvn":"scvn","standardcharteredvn":"scvn",
    # IBK
    "ibk":"ibk","congnghiephanquoc":"ibk",
    # Indovina
    "indovinabank":"ivb","ivb":"ivb","indovina":"ivb",
    # VRB
    "vrb":"vrb","vietnga":"vrb",
    # Nonghyup
    "nonghyup":"nonghyup",
    # DongA Bank
    "dongabank":"dab","dab":"dab","donga":"dab","dongasean":"dab",
    # PGBank
    "pgbank":"pgb","pgb":"pgb","petrolimex":"pgb","xangdau":"pgb",
    # SaigonBank (SGBL) — alias thực tế: svb
    "saigonbanksgbl":"sgbl","sgbl":"sgbl","saigoncongth":"sgbl","saigonbk":"sgbl",
    # SVB = ShinhanBank (thực tế alias dùng phổ biến)
    "svb": "shbvn",
    # VietCapitalBank
    "vietcapitalbank":"vcb2","bancviet":"vcb2","banviet":"vcb2",
    # ViettelMoney
    "viettelmoney":"viettelm","viettelm":"viettelm",
    # VNPTMoney
    "vnptmoney":"vnptm","vnptm":"vnptm",
    # LioBank
    "liobank":"liobank","lio":"liobank",
    # Ubank
    "ubank":"ubank",
    # BIDC
    "bidc":"bidc",
    # Cake by VPBank
    "cake":"cake","cakebyvpbank":"cake",
    # Timo
    "timo":"timo",
    # CoopBank
    "coopbank":"coopbank","coop":"coopbank","hoptacxa":"coopbank",
    # UnitedOverseas
    "unitedoverseas":"uob","uob":"uob",
    # CathayUnitedBank
    "cathayunitedbank":"cathay","cathay":"cathay",
    # KEBHana
    "kebhana":"kebhanabn","kebhanabn":"kebhanabn","kebhanahcm":"kebhanahcm",
    # Kookmin
    "kookminho":"kookminho","kookminhn":"kookminhn","kookminhcm":"kookminhcm",
    # DBSBank
    "dbsbank":"dbs","dbs":"dbs",
    # CitibankHN
    "citibank":"citibank","citibankhn":"citibank",
    # BNPHN/BNPHCM
    "bnp":"bnphn","bnphn":"bnphn","bnphcm":"bnphcm",
}

def _vi_normalize(text: str) -> str:
    """Chuẩn hoá tên có dấu tiếng Việt về ASCII để tra cứu."""
    import unicodedata
    nfd = unicodedata.normalize("NFD", text)
    ascii_str = "".join(c for c in nfd if unicodedata.category(c) != "Mn")
    return re.sub(r"[^a-z0-9]", "", ascii_str.lower())


def resolve_bank_code(raw: str) -> str:
    key = re.sub(r"[^a-z0-9]", "", raw.lower())
    if key in BANK_ALIAS:
        return BANK_ALIAS[key]
    vi_key = _vi_normalize(raw)
    return BANK_ALIAS.get(vi_key, key)


def looks_like_bank(line: str) -> bool:
    # Key 1: strip non-alphanum
    key = re.sub(r"[^a-z0-9]", "", line.lower())
    if key in BANK_ALIAS:
        return True
    # Key 2: normalize dấu tiếng Việt (VIỆT Á → vieta → vab check)
    vi_key = _vi_normalize(line)
    if vi_key in BANK_ALIAS:
        return True
    # Key 3: chỉ gồm chữ Latin (mã ngân hàng viết tắt)
    if re.fullmatch(r"[a-zA-Z]{2,16}", line.strip()):
        return True
    return False


def digits_only(text: str) -> str:
    return re.sub(r"\D", "", text)


def parse_amount(text: str) -> Optional[int]:
    d = digits_only(text)
    return int(d) if d and int(d) >= 1_000 else None


def is_clearly_money(line: str) -> bool:
    """Dòng tiền RÕ RÀNG: có dấu phẩy/chấm hoặc chứa đ."""
    return bool(re.search(r"[,.]|đ", line)) and parse_amount(line) is not None


def format_money(amount: int) -> str:
    return f"{amount:,}"


def looks_like_account(line: str) -> bool:
    compact = re.sub(r"[\s\-.]", "", line)
    return compact.isdigit() and 6 <= len(compact) <= 20


def clean_line(line: str) -> str:
    # Strip non-breaking space \xa0 và zero-width chars copy từ app
    line = re.sub(r"[\xa0\u200b\u200c\u200d\ufeff]+", " ", line)
    return line.strip().strip('"').strip("'").strip("\u201c").strip("\u201d").strip()


def is_separator(line: str) -> bool:
    s = line.strip()
    if not s:
        return False
    return bool(re.fullmatch(r"[\-_=\u2014\u2013\.\s>»<~*#]{2,}", s))


def _split_bank_stk(line: str) -> List[str]:
    """Tách dòng 'BankName – 0123456789' thành 2 dòng riêng."""
    m = re.match(r"^(.+?)\s*[\u2013\u2014\-:|]\s*(\d[\d\s]{5,19})$", line.strip())
    if m:
        left = m.group(1).strip()
        right = re.sub(r"\s", "", m.group(2)).strip()
        if looks_like_bank(left):
            return [left, right]
    return [line]


def normalize_lines(text: str) -> List[str]:
    lines = []
    for raw in text.replace("\r", "\n").split("\n"):
        line = clean_line(raw)
        if not line or is_separator(line):
            continue
        # Chuẩn hoá tab / nhiều space
        line = re.sub(r"[\t]+ ", " ", line).strip()
        for part in _split_bank_stk(line):
            if part:
                lines.append(part)
    return lines


def _in_bank_alias(line: str) -> bool:
    key = re.sub(r"[^a-z0-9]", "", line.lower())
    return key in BANK_ALIAS or _vi_normalize(line) in BANK_ALIAS


def _looks_like_bank_code(line: str) -> bool:
    """Bank code viết tắt không dấu: tcb, ACB, MSB, tpb..."""
    return bool(re.fullmatch(r"[a-zA-Z]{2,10}", line.strip()))


def parse_two_info(name_line: str, bank_stk_line: str, label: str) -> Dict[str, str]:
    """
    Parse format 2 dòng:
      Dòng 1: tên tài khoản
      Dòng 2: bank + STK (dãy số dài nhất trong dòng = STK, phần còn lại = bank)

    Không phân biệt dấu gạch, khoảng trắng hay ký tự đặc biệt.
    """
    name = name_line.strip()
    if not name:
        raise ValueError(f"{label}: tên tài khoản trống")

    # Tìm dãy số dài nhất trong dòng → đó là STK
    matches = re.findall(r'\d+', bank_stk_line)
    if not matches:
        raise ValueError(f"{label}: không tìm được STK trong '{bank_stk_line}'")

    stk_match = max(matches, key=len)
    if len(stk_match) < 6:
        raise ValueError(f"{label}: STK quá ngắn trong '{bank_stk_line}'")

    account = stk_match
    # Bank = bỏ STK ra, dọn ký tự thừa
    bank = re.sub(re.escape(stk_match), '', bank_stk_line)
    bank = re.sub(r'[\-|/\s]+', ' ', bank).strip()

    if not bank:
        raise ValueError(f"{label}: không tìm được tên bank trong '{bank_stk_line}'")

    return {"bank": bank, "account": account, "name": name}


def parse_three_info(lines: List[str], label: str) -> Dict[str, str]:
    """
    Nhận diện bank / STK / tên trong 3 dòng (thứ tự bất kỳ).
    Priority:
    1. STK: dòng toàn số 6-20 ký tự
    2. Bank: dòng trong BANK_ALIAS → bank code ngắn → dòng ngắn hơn
    3. Tên: dòng còn lại
    """
    if len(lines) != 3:
        raise ValueError(f"{label}: cần 3 dòng, có {len(lines)}: {lines}")

    account = bank = name = None

    # Bước 1: tìm STK
    for l in lines:
        if looks_like_account(l) and not account:
            account = digits_only(l)

    remaining = [l for l in lines if not (account and digits_only(l) == account)]

    # Bước 2: tìm bank trong 2 dòng còn lại
    if len(remaining) == 2:
        a, b = remaining
        a_alias = _in_bank_alias(a)
        b_alias = _in_bank_alias(b)
        a_code  = _looks_like_bank_code(a)
        b_code  = _looks_like_bank_code(b)

        if a_alias and not b_alias:
            bank, name = a, b
        elif b_alias and not a_alias:
            bank, name = b, a
        elif a_code and not b_code:
            bank, name = a, b
        elif b_code and not a_code:
            bank, name = b, a
        else:
            # Không rõ → dòng ngắn hơn là bank
            bank, name = (a, b) if len(a) <= len(b) else (b, a)
    elif len(remaining) == 1:
        # STK không nhận diện được → fallback
        bank = remaining[0]
        name = ""
    
    if not account:
        raise ValueError(f"Không nhận diện được STK trong {label}: {lines}")
    if not bank:
        raise ValueError(f"Không nhận diện được ngân hàng trong {label}: {lines}")
    if not name:
        raise ValueError(f"Không nhận diện được tên tài khoản trong {label}: {lines}")

    return {"bank": bank.strip(), "account": account, "name": name.strip()}
def _raw_blocks_from_text(text: str) -> List[List[str]]:
    """
    Tách text thành các block theo dòng trống / separator.
    Sau khi tách, nếu block chỉ có 1 dòng là số tiền → gộp vào block trước.
    Xử lý trường hợp user nhập dòng trống giữa bank-STK và số tiền.
    """
    blocks: List[List[str]] = []
    current: List[str] = []

    for raw in text.replace("\r", "\n").split("\n"):
        line = clean_line(raw)
        if not line or is_separator(line):
            if current:
                blocks.append(current)
                current = []
        else:
            current.append(line)

    if current:
        blocks.append(current)

    # Gộp block chỉ có số tiền vào block trước
    merged: List[List[str]] = []
    for blk in blocks:
        if (len(blk) == 1
                and parse_amount(blk[0]) is not None
                and merged):
            merged[-1].append(blk[0])
        else:
            merged.append(blk)

    return merged


def _parse_block_fixed(lines: List[str], label: str, require_amount: bool = True) -> Dict[str, Any]:
    """
    Parse cụm — hỗ trợ cả 2 format:

    Format 2 dòng (mới):
      dòng 1 : tên tài khoản
      dòng 2 : bank - STK
      dòng 3 : số tiền
      dòng 4 : nội dung (tuỳ chọn)

    Format 3 dòng (cũ):
      dòng 1-3 : thứ tự bất kỳ trong {bank, STK, tên}
      dòng 4   : số tiền
      dòng 5   : nội dung (tuỳ chọn)

    Detect tự động: nếu dòng 1 không có STK và dòng 2 có STK → format 2 dòng.
    """
    if len(lines) < 2:
        raise ValueError(f"{label}: cụm phải có ít nhất 2 dòng, nhận {len(lines)}: {lines}")

    content = DEFAULT_TRANSFER_CONTENT
    amount  = None
    result  = None

    # ── Detect format ────────────────────────────────────────────────────────
    # Format 2 dòng chuẩn: dòng 1=tên, dòng 2=bank-STK
    is_2line = (
        len(lines) >= 2
        and not _line_contains_stk(lines[0])
        and _is_bank_stk_line(lines[1])
    )

    # Format 2 dòng biến thể: tên / bank / STK (bank và STK tách riêng)
    is_2line_split = (
        not is_2line
        and len(lines) >= 3
        and not _line_contains_stk(lines[0])
        and not _line_contains_stk(lines[1])
        and bool(re.search(r'[a-zA-ZÀ-ỹ]', lines[1]))
        and bool(re.match(r'^\d{6,20}$', lines[2].strip()))
    )

    if is_2line:
        result = parse_two_info(lines[0], lines[1], label)
        rest   = lines[2:]
    elif is_2line_split:
        merged = f"{lines[1].strip()} {lines[2].strip()}"
        result = parse_two_info(lines[0], merged, label)
        rest   = lines[3:]
    else:
        # Format 3 dòng cũ
        if len(lines) < 3:
            raise ValueError(f"{label}: cụm phải có ít nhất 3 dòng, nhận {len(lines)}: {lines}")
        result = parse_three_info(lines[:3], label)
        rest   = lines[3:]

    # ── Đọc số tiền + nội dung từ phần còn lại ───────────────────────────────
    if rest:
        amount = parse_amount(rest[0])
        if amount is None and require_amount:
            raise ValueError(
                f"{label}: dòng số tiền '{rest[0]}' không hợp lệ.\n"
                f"Số tiền ghi dạng: 93,493,000"
            )
        if len(rest) >= 2 and rest[1]:
            content = rest[1]
    elif require_amount:
        raise ValueError(f"{label}: thiếu dòng số tiền.")

    if amount is not None:
        result["amount"] = amount
    result["content"] = content
    return result


def _line_contains_stk(line: str) -> bool:
    """Dòng có chứa dãy số ≥ 6 ký tự liên tiếp = có STK."""
    return bool(re.search(r'\d{6,}', line))


def _is_bank_stk_line(line: str) -> bool:
    """
    Dòng chứa cả bank lẫn STK — phân biệt với STK thuần số.
    Điều kiện: có dãy số ≥ 6 ký tự VÀ có chữ cái (tên bank).
    """
    has_stk  = bool(re.search(r'\d{6,}', line))
    has_text = bool(re.search(r'[a-zA-ZÀ-ỹ]', line))
    return has_stk and has_text


def _extract_sender_lines(block: List[str]) -> tuple:
    """
    Trích xuất thông tin người chuyển/nhận từ block.
    Hỗ trợ theo thứ tự ưu tiên:
      1. Format 2 dòng: dòng 1=tên, dòng 2=bank-STK (kể cả dính nhau: ncb-100812677734)
      2. Format 1 dòng gộp: "TCB - 3346907593 - Bùi Đình Kiên"
      3. Format 3 dòng riêng (thứ tự bất kỳ)

    Trả về: (parsed_dict, leftover) hoặc (None, block)
    """
    # Thử format 2 dòng: dòng 1 không chứa STK, dòng 2 là bank-STK line
    if len(block) >= 2:
        line1, line2 = block[0], block[1]
        if not _line_contains_stk(line1) and _is_bank_stk_line(line2):
            try:
                parsed = parse_two_info(line1, line2, "block")
                return parsed, block[2:]
            except ValueError:
                pass

    # Thử format 2 dòng biến thể: bank và STK trên 2 dòng riêng
    # Ví dụ: "Phan Thị Thùy Linh" / "NCB -" / "100812677734"
    # → gộp dòng 2+3 thành "NCB - 100812677734" rồi parse
    if len(block) >= 3:
        line1, line2, line3 = block[0], block[1], block[2]
        line2_has_text = bool(re.search(r'[a-zA-ZÀ-ỹ]', line2))
        line3_is_stk   = bool(re.match(r'^\d{6,20}$', line3.strip()))
        if (not _line_contains_stk(line1)
                and not _line_contains_stk(line2) and line2_has_text
                and line3_is_stk):
            merged_line = f"{line2.strip()} {line3.strip()}"
            try:
                parsed = parse_two_info(line1, merged_line, "block")
                return parsed, block[3:]
            except ValueError:
                pass

    # Thử 1 dòng gộp — tách theo dấu - | /
    if len(block) >= 1:
        first = block[0]
        parts = re.split(r"\s*[-|/]\s+|\s+[-|/]\s*", first)
        parts = [p.strip() for p in parts if p.strip()]
        if len(parts) == 3:
            has_stk = any(looks_like_account(p) for p in parts)
            if has_stk:
                try:
                    info = parse_three_info(parts, "block")
                    return info, block[1:]
                except ValueError:
                    pass

    # Thử 3 dòng riêng
    if len(block) >= 3:
        candidate = block[:3]
        has_stk  = any(_line_contains_stk(l) for l in candidate)
        has_bank = any(_in_bank_alias(l) or _looks_like_bank_code(l) for l in candidate)
        if has_stk and has_bank:
            try:
                info = parse_three_info(candidate, "block")
                return info, block[3:]
            except ValueError:
                pass
        if has_stk:
            try:
                info = parse_three_info(candidate, "block")
                return info, block[3:]
            except ValueError:
                pass

    return None, block


def _block_has_amount(block: List[str]) -> bool:
    """
    Kiểm tra block có chứa dòng số tiền không.
    Dùng is_clearly_money (phải có dấu phẩy/chấm hoặc đ) để tránh
    nhầm STK thuần số (100812677734) với số tiền.
    Kiểm tra từ index 2 trở đi để bao cả format 2 dòng lẫn 3 dòng.
    """
    for line in block[2:]:
        if is_clearly_money(line):
            return True
    return False


def _merge_receiver_blocks(raw_blocks: List[List[str]]) -> List[List[str]]:
    """
    Các block đã được tách sẵn bởi dòng trống — pass thẳng qua.
    Nhưng nếu block chỉ có 2 dòng [tên, bank] và block tiếp theo
    chỉ có [STK, tiền] thì KHÔNG gộp — đã được xử lý bởi _parse_block_fixed.
    """
    return [blk for blk in raw_blocks if blk]


def parse_order_form(text: str) -> Dict[str, Any]:
    """
    Parse form hỗ trợ 3 case:

    Case 1 & 3 — 1 người chuyển → 1 hoặc nhiều người nhận:
      Block ĐẦU không có số tiền  = người chuyển
      Các block còn lại có tiền   = người nhận
      QR được tạo cho từng người nhận.

    Case 2 — nhiều người chuyển → 1 người nhận:
      Block CUỐI không có số tiền = người nhận
      Các block trước có tiền     = từng người chuyển
      QR được tạo cho từng người chuyển (QR hiển thị TK người nhận).

    Detect tự động dựa vào vị trí block không có tiền.
    """
    raw_blocks = _raw_blocks_from_text(text)

    if not raw_blocks:
        raise ValueError("Form trống.")

    if len(raw_blocks) < 2:
        raise ValueError(
            "Form cần ít nhất 2 block (người chuyển và người nhận).\n"
            "Tách các block bằng dòng trống."
        )

    # ── Detect case ──────────────────────────────────────────────────────────
    first_has_amount = _block_has_amount(raw_blocks[0])
    last_has_amount  = _block_has_amount(raw_blocks[-1])

    # Case 2: block đầu có tiền, block cuối không có tiền
    is_case2 = first_has_amount and not last_has_amount

    # ── Parse Case 2: nhiều chuyển → 1 nhận ──────────────────────────────────
    if is_case2:
        # Block cuối = người nhận duy nhất
        receiver_block = raw_blocks[-1]
        receiver_parsed, _ = _extract_sender_lines(receiver_block)
        if receiver_parsed is None:
            raise ValueError(
                "Không nhận diện được thông tin người nhận (block cuối).\n"
                "Cần có: tên ngân hàng, số tài khoản, tên tài khoản."
            )
        receiver_info = receiver_parsed

        # Các block trước = danh sách người chuyển (mỗi block có tiền)
        sender_blocks_raw = raw_blocks[:-1]
        sender_blocks = _merge_receiver_blocks(sender_blocks_raw)

        if not sender_blocks:
            raise ValueError("Không tìm thấy thông tin người chuyển.")

        senders = []
        for i, blk in enumerate(sender_blocks, 1):
            s = _parse_block_fixed(blk, f"người chuyển {i}", require_amount=True)
            s["index"] = i
            senders.append(s)

        # Build parsed — đảo vai trò: QR tạo cho người nhận, sender là list
        # Để tương thích với send_order_qrs, map sang cùng structure:
        # "sender"    = người nhận (TK nhận tiền, dùng để tạo QR)
        # "receivers" = danh sách người chuyển (mỗi người 1 QR)
        # "case"      = 2 (để caption hiển thị đúng chiều)
        return {
            "case": 2,
            "sender": receiver_info,          # TK nhận — dùng tạo QR
            "receivers": senders,             # Danh sách người chuyển
            "total_amount": sum(s["amount"] for s in senders),
            "total_qr": len(senders),
        }

    # ── Parse Case 1 & 3: 1 chuyển → nhiều nhận ──────────────────────────────
    sender_block = raw_blocks[0]
    sender_parsed, leftover = _extract_sender_lines(sender_block)
    if sender_parsed is None:
        raise ValueError(
            "Không nhận diện được thông tin người chuyển.\n"
            "Cần có: tên ngân hàng, số tài khoản, tên tài khoản."
        )
    sender = sender_parsed

    receiver_blocks_raw = raw_blocks[1:]

    if not receiver_blocks_raw:
        # Không có dòng trống tách block → dùng phần dư sau sender
        body_lines = leftover
        if not body_lines:
            raise ValueError("Không tìm thấy thông tin người nhận.")
        receiver_blocks = [body_lines[i:i+4] for i in range(0, len(body_lines), 4)]
    else:
        receiver_blocks = _merge_receiver_blocks(receiver_blocks_raw)

    if not receiver_blocks:
        raise ValueError("Không tìm thấy thông tin người nhận.")

    receivers = []
    for i, blk in enumerate(receiver_blocks, 1):
        rec = _parse_block_fixed(blk, f"người nhận đơn {i}", require_amount=True)
        rec["index"] = i
        receivers.append(rec)

    return {
        "case": 1,
        "sender": sender,
        "receivers": receivers,
        "total_amount": sum(r["amount"] for r in receivers),
        "total_qr": len(receivers),
    }


def generate_order_code() -> str:
    """
    Format: XT + DDHHMMSS[+seq nếu trùng].
    An toàn với concurrent requests nhờ DB UNIQUE constraint.
    """
    base = "XT" + now_local().strftime("%d%H%M%S")
    conn = db_connect()
    cur = conn.cursor()
    candidate = base
    n = 1
    while True:
        cur.execute("SELECT 1 FROM orders WHERE order_code=?", (candidate,))
        if not cur.fetchone():
            conn.close()
            return candidate
        candidate = f"{base}{n:02d}"
        n += 1


def sender_display_name(message: Message) -> str:
    user = message.from_user
    if not user:
        return ""
    if user.username:
        return f"@{user.username}"
    return " ".join(x for x in [user.first_name, user.last_name] if x) or str(user.id)


def chat_title(message: Message) -> str:
    chat = message.chat
    return getattr(chat, "title", None) or getattr(chat, "full_name", None) or str(chat.id)


def safe_filename(text: str) -> str:
    text = re.sub(r'[\\/:*?"<>|]+', " ", text)
    return re.sub(r"\s+", " ", text).strip()[:120]


# ─────────────────────────── ACTIVATION GUARD ────────────────────────────────

def chat_is_active(chat_id: int) -> bool:
    conn = db_connect()
    cur = conn.cursor()
    cur.execute("SELECT 1 FROM activated_chats WHERE chat_id=?", (chat_id,))
    ok = cur.fetchone() is not None
    conn.close()
    return ok


# ─────────────────────────── PERMISSION ──────────────────────────────────────

async def is_bot_admin(chat_id: int, user_id: int) -> bool:
    """Check global_admins — chat_id không còn dùng, giữ signature tương thích."""
    if user_id in SUPER_ADMIN_IDS:
        return True
    conn = db_connect()
    cur = conn.cursor()
    cur.execute("SELECT 1 FROM global_admins WHERE user_id=?", (user_id,))
    ok = cur.fetchone() is not None
    conn.close()
    return ok

def is_superadmin(user_id: int) -> bool:
    return user_id in SUPER_ADMIN_IDS


# ─────────────────────────── QR GENERATION ───────────────────────────────────

async def download_vietqr_image(
    bank: str, account: str, account_name: str,
    amount: int, content: str, output_path: str
) -> None:
    bank_code = resolve_bank_code(bank)
    query = urlencode({"amount": amount, "addInfo": content, "accountName": account_name})
    url = f"https://img.vietqr.io/image/{bank_code}-{account}-compact2.png?{query}"

    async with aiohttp.ClientSession() as session:
        async with session.get(url, timeout=aiohttp.ClientTimeout(total=30)) as resp:
            if resp.status != 200:
                raise RuntimeError(
                    f"Ngân hàng <b>{bank}</b> hoặc STK <code>{account}</code> không hợp lệ "
                    f"(HTTP {resp.status}). Kiểm tra lại mã ngân hàng và số tài khoản."
                )
            data = await resp.read()

    # Kiểm tra bytes trả về có phải PNG hợp lệ không
    if not data.startswith(b"\x89PNG"):
        raise RuntimeError(
            f"VietQR trả về dữ liệu không phải ảnh cho ngân hàng <b>{bank}</b> "
            f"STK <code>{account}</code>. Kiểm tra lại mã ngân hàng."
        )
    Path(output_path).write_bytes(data)


def make_thumbnail(image_path: str, thumb_path: str) -> None:
    try:
        with Image.open(image_path) as img:
            img.thumbnail((180, 180))
            img.save(thumb_path, "PNG", optimize=True)
    except Exception:
        # Fallback: tạo thumbnail trắng 1x1 để không crash
        Image.new("RGB", (1, 1), "white").save(thumb_path, "PNG")


# ─────────────────────────── CAPTION ─────────────────────────────────────────

def build_caption(order_code: str, parsed: Dict[str, Any]) -> str:
    s = parsed["sender"]
    total = parsed["total_qr"]
    lines = [
        f"✅ <b>THÔNG TIN NGƯỜI CHUYỂN</b>",
        f"<b>{s['name']}</b>  —  {s['bank'].upper()}  —  <code>{s['account']}</code>",
        "",
        f"📦 Mã đơn: <b><code>{order_code}</code></b>  |  💰 <code>{format_money(parsed['total_amount'])}</code>  ({total} QR)",
        "",
        f"🏦 <b>THÔNG TIN NGƯỜI NHẬN</b>",
    ]
    for r in parsed["receivers"]:
        lines += [
            f"",
            f"<b>——— Đơn {r['index']}/{total} ———</b>",
            f"👤 Tên TK: <b>{r['name']}</b>",
            f"🏦 Ngân hàng: <b>{r['bank'].upper()}</b>  —  <code>{r['account']}</code>",
            f"💵 Số tiền: <code>{format_money(r['amount'])}</code>",
            *([ f"📝 Nội dung: {r['content']}" ] if r.get("content") else []),
        ]
    return "\n".join(lines)


def build_caption_single(order_code: str, parsed: Dict[str, Any], r: Dict[str, Any]) -> str:
    """
    Caption riêng cho từng QR — hỗ trợ cả Case 1/3 và Case 2.

    Case 1/3 (1 chuyển → nhiều nhận):
      Hiển thị: người chuyển cố định / người nhận theo từng QR

    Case 2 (nhiều chuyển → 1 nhận):
      Hiển thị: người chuyển theo từng QR / người nhận cố định
      parsed["sender"] = TK nhận (dùng tạo QR)
      parsed["receivers"][i] = người chuyển thứ i
    """
    total = parsed["total_qr"]
    case  = parsed.get("case", 1)

    if case == 2:
        # r = người chuyển, parsed["sender"] = người nhận
        receiver = parsed["sender"]
        lines = [
            f"👤 <b>NGƯỜI CHUYỂN — {r['index']}/{total}</b>",
            f"<b>{r['name']}</b>  —  {r['bank'].upper()}  —  <code>{r['account']}</code>",
            f"💵 Số tiền: <code>{format_money(r['amount'])}</code>",
            *([ f"📝 Nội dung: {r['content']}" ] if r.get("content") else []),
            "",
            f"📦 Mã đơn: <b><code>{order_code}</code></b>  |  💰 <code>{format_money(parsed['total_amount'])}</code>  ({total} QR)",
            "",
            f"🏦 <b>NGƯỜI NHẬN</b>",
            f"<b>{receiver['name']}</b>  —  {receiver['bank'].upper()}  —  <code>{receiver['account']}</code>",
        ]
    else:
        # r = người nhận, parsed["sender"] = người chuyển
        s = parsed["sender"]
        lines = [
            f"✅ <b>THÔNG TIN NGƯỜI CHUYỂN</b>",
            f"<b>{s['name']}</b>  —  {s['bank'].upper()}  —  <code>{s['account']}</code>",
            "",
            f"📦 Mã đơn: <b><code>{order_code}</code></b>  |  💰 <code>{format_money(parsed['total_amount'])}</code>  ({total} QR)",
            "",
            f"🏦 <b>THÔNG TIN NGƯỜI NHẬN</b>",
            "",
            f"<b>——— Đơn {r['index']}/{total} ———</b>",
            f"👤 Tên TK: <b>{r['name']}</b>",
            f"🏦 Ngân hàng: <b>{r['bank'].upper()}</b>  —  <code>{r['account']}</code>",
            f"💵 Số tiền: <code>{format_money(r['amount'])}</code>",
            *([ f"📝 Nội dung: {r['content']}" ] if r.get("content") else []),
        ]
    return "\n".join(lines)


# ─────────────────────────── SEND QRs ────────────────────────────────────────

def resize_qr(src_path: str, dst_path: str, size: int = 280) -> None:
    """
    Resize QR về 280x280 JPEG ~12-18KB.
    Đủ nhỏ để preview nhanh, vẫn scan được bằng điện thoại.
    """
    with Image.open(src_path) as img:
        bg = Image.new("RGB", img.size, "white")
        if img.mode in ("RGBA", "LA", "P"):
            try:
                bg.paste(img.convert("RGBA"), mask=img.convert("RGBA").split()[3])
            except Exception:
                bg.paste(img.convert("RGB"))
        else:
            bg.paste(img.convert("RGB"))
        bg = bg.resize((size, size), Image.LANCZOS)
        bg.save(dst_path, "JPEG", quality=75, optimize=True)


async def _make_qr_bytes(raw_path: str) -> tuple:
    """Trả về (doc_bytes PNG, thumb_bytes JPEG 90x90)."""
    import io
    with Image.open(raw_path) as img:
        bg = Image.new("RGB", img.size, "white")
        if img.mode in ("RGBA", "LA", "P"):
            try:
                bg.paste(img.convert("RGBA"), mask=img.convert("RGBA").split()[3])
            except Exception:
                bg.paste(img.convert("RGB"))
        else:
            bg.paste(img.convert("RGB"))
        doc_io = io.BytesIO()
        bg.save(doc_io, format="PNG", optimize=True)
        thumb = bg.copy()
        thumb.thumbnail((90, 90), Image.LANCZOS)
        thumb_io = io.BytesIO()
        thumb.save(thumb_io, format="JPEG", quality=70)
    return doc_io.getvalue(), thumb_io.getvalue()


async def send_order_qrs(
    bot: Bot, message: Message, order_code: str, parsed: Dict[str, Any]
) -> List[int]:
    """
    Gửi từng QR thành 1 tin nhắn riêng — hỗ trợ cả 3 case:

    Case 1/3 (1 chuyển → nhiều nhận):
      QR tạo theo TK người nhận + số tiền người nhận
      Tên file: {tên người nhận} {bank} - {số tiền}.png

    Case 2 (nhiều chuyển → 1 nhận):
      QR tạo theo TK người nhận (parsed["sender"]) + số tiền từng người chuyển
      Tên file: {tên người chuyển} -> {tên người nhận} - {số tiền}.png
    """
    from aiogram.types import BufferedInputFile

    case      = parsed.get("case", 1)
    receivers = parsed["receivers"]   # Case 1/3: người nhận | Case 2: người chuyển
    sent_ids: List[int] = []

    with tempfile.TemporaryDirectory() as tmpdir:
        # ── Build danh sách tham số download QR ──────────────────────────────
        if case == 2:
            # QR hướng về TK người nhận, amount từ từng người chuyển
            recv_info = parsed["sender"]   # TK nhận tiền
            download_params = [
                dict(
                    bank=recv_info["bank"],
                    account=recv_info["account"],
                    account_name=recv_info["name"],
                    amount=r["amount"],
                    content=r.get("content") or DEFAULT_TRANSFER_CONTENT,
                    output_path=os.path.join(tmpdir, f"{order_code}_{r['index']}_raw.png"),
                )
                for r in receivers
            ]
        else:
            download_params = [
                dict(
                    bank=r["bank"],
                    account=r["account"],
                    account_name=r["name"],
                    amount=r["amount"],
                    content=r["content"],
                    output_path=os.path.join(tmpdir, f"{order_code}_{r['index']}_raw.png"),
                )
                for r in receivers
            ]

        raw_paths = [p["output_path"] for p in download_params]

        # Download song song
        await asyncio.gather(*[
            download_vietqr_image(**{k: v for k, v in p.items() if k != "output_path"},
                                  output_path=p["output_path"])
            for p in download_params
        ])

        # Gửi từng QR thành 1 tin nhắn riêng — tuần tự để đúng thứ tự
        first_msg_id = None
        for i, (r, raw_p) in enumerate(zip(receivers, raw_paths)):
            doc_bytes, thumb_bytes = await asyncio.get_event_loop().run_in_executor(
                None, lambda rp=raw_p: _make_qr_bytes_sync(rp)
            )

            if case == 2:
                recv_info = parsed["sender"]
                fname = safe_filename(
                    f"{r['name']} -> {recv_info['name']} {recv_info['bank'].upper()} - {format_money(r['amount'])}.png"
                )
            else:
                fname = safe_filename(
                    f"{r['name']} {r['bank'].upper()} - {format_money(r['amount'])}.png"
                )

            caption   = build_caption_single(order_code, parsed, r)
            doc_buf   = BufferedInputFile(doc_bytes,   filename=fname)
            thumb_buf = BufferedInputFile(thumb_bytes, filename="t.jpg")

            # QR đầu tiên reply vào form gốc; các QR sau reply vào QR trước đó
            reply_to = message.message_id if i == 0 else sent_ids[-1]

            msg = await bot.send_document(
                chat_id=message.chat.id,
                document=doc_buf,
                thumbnail=thumb_buf,
                caption=caption,
                parse_mode=ParseMode.HTML,
                reply_to_message_id=reply_to,
            )
            sent_ids.append(msg.message_id)
            if i == 0:
                first_msg_id = msg.message_id

        # Tin nhắn cuối: inline buttons reply vào QR đầu tiên
        from aiogram.types import InlineKeyboardMarkup, InlineKeyboardButton
        kb = InlineKeyboardMarkup(inline_keyboard=[[
            InlineKeyboardButton(text="✅ Đã gửi đơn", callback_data=f"sent:{order_code}"),
            InlineKeyboardButton(text="❌ Hủy đơn",   callback_data=f"cancel:{order_code}"),
        ]])
        btn_msg = await bot.send_message(
            chat_id=message.chat.id,
            text=f"📦 <b><code>{order_code}</code></b>  |  {parsed['total_qr']} QR  |  <code>{format_money(parsed['total_amount'])}</code>",
            parse_mode=ParseMode.HTML,
            reply_markup=kb,
            reply_to_message_id=first_msg_id,
        )
        sent_ids.append(btn_msg.message_id)

    return sent_ids


def _make_qr_bytes_sync(raw_path: str) -> tuple:
    """Sync version cho run_in_executor."""
    import io
    with Image.open(raw_path) as img:
        bg = Image.new("RGB", img.size, "white")
        if img.mode in ("RGBA", "LA", "P"):
            try:
                bg.paste(img.convert("RGBA"), mask=img.convert("RGBA").split()[3])
            except Exception:
                bg.paste(img.convert("RGB"))
        else:
            bg.paste(img.convert("RGB"))
        doc_io = io.BytesIO()
        bg.save(doc_io, format="PNG", optimize=True)
        thumb = bg.copy()
        thumb.thumbnail((90, 90), Image.LANCZOS)
        thumb_io = io.BytesIO()
        thumb.save(thumb_io, format="JPEG", quality=70)
    return doc_io.getvalue(), thumb_io.getvalue()


# ─────────────────────────── DB SAVE / QUERY ─────────────────────────────────

def save_order_to_db(
    order_code: str, message: Message,
    parsed: Dict[str, Any], sent_ids: List[int]
) -> None:
    conn = db_connect()
    cur = conn.cursor()
    ts  = now_local().strftime("%Y-%m-%d %H:%M:%S")
    case = parsed.get("case", 1)

    if case == 2:
        # Case 2: parsed["sender"] = TK nhận tiền (block cuối)
        #         parsed["receivers"] = danh sách người chuyển
        # Lưu DB đúng vai trò:
        #   orders.sender_* = người chuyển đầu tiên (đại diện)
        #   receivers = từng người chuyển (bank/account/name/amount)
        # Nhưng để Excel/báo cáo đọc được,
        # dùng cột sender = TK NHẬN (block cuối) như là "đầu kia"
        # và receivers = từng người CHUYỂN
        recv_tk = parsed["sender"]          # TK nhận thực tế
        senders = parsed["receivers"]       # Danh sách người chuyển

        cur.execute("""
            INSERT INTO orders
                (order_code, created_at, chat_id, group_name,
                 creator_user_id, creator_name, creator_username,
                 creator_message_id, qr_message_id, button_message_id,
                 sender_bank, sender_account, sender_name,
                 total_amount, total_qr, status)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,'active')
        """, (
            order_code, ts, message.chat.id, chat_title(message),
            message.from_user.id, sender_display_name(message),
            f"@{message.from_user.username}" if message.from_user.username else sender_display_name(message),
            message.message_id,
            sent_ids[0] if sent_ids else None,
            sent_ids[-1] if sent_ids else None,
            recv_tk["bank"], recv_tk["account"], recv_tk["name"],   # TK nhận
            parsed["total_amount"], parsed["total_qr"],
        ))

        for r, mid in zip(senders, sent_ids):
            cur.execute("""
                INSERT INTO receivers
                    (order_code, receiver_index, receiver_bank,
                     receiver_account, receiver_name, amount, content, message_id)
                VALUES (?,?,?,?,?,?,?,?)
            """, (order_code, r["index"], r["bank"],
                  r["account"], r["name"], r["amount"], r.get("content",""), mid))
            cur.execute("""
                INSERT OR REPLACE INTO order_messages (chat_id, message_id, order_code)
                VALUES (?,?,?)
            """, (message.chat.id, mid, order_code))

    else:
        # Case 1/3: logic cũ
        s = parsed["sender"]
        cur.execute("""
            INSERT INTO orders
                (order_code, created_at, chat_id, group_name,
                 creator_user_id, creator_name, creator_username,
                 creator_message_id, qr_message_id, button_message_id,
                 sender_bank, sender_account, sender_name,
                 total_amount, total_qr, status)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,'active')
        """, (
            order_code, ts, message.chat.id, chat_title(message),
            message.from_user.id, sender_display_name(message),
            f"@{message.from_user.username}" if message.from_user.username else sender_display_name(message),
            message.message_id,
            sent_ids[0] if sent_ids else None,
            sent_ids[-1] if sent_ids else None,
            s["bank"], s["account"], s["name"],
            parsed["total_amount"], parsed["total_qr"],
        ))

        for r, mid in zip(parsed["receivers"], sent_ids):
            cur.execute("""
                INSERT INTO receivers
                    (order_code, receiver_index, receiver_bank,
                     receiver_account, receiver_name, amount, content, message_id)
                VALUES (?,?,?,?,?,?,?,?)
            """, (order_code, r["index"], r["bank"],
                  r["account"], r["name"], r["amount"], r["content"], mid))
            cur.execute("""
                INSERT OR REPLACE INTO order_messages (chat_id, message_id, order_code)
                VALUES (?,?,?)
            """, (message.chat.id, mid, order_code))

    conn.commit()
    conn.close()


async def get_order_by_message(chat_id: int, message_id: int) -> Optional[str]:
    """Tìm order_code qua message_id — check cả QR bot gửi lẫn form gốc người dùng."""
    conn = db_connect()
    cur = conn.cursor()
    # Tìm trong QR messages (bot gửi)
    cur.execute(
        "SELECT order_code FROM order_messages WHERE chat_id=? AND message_id=?",
        (chat_id, message_id)
    )
    row = cur.fetchone()
    if row:
        conn.close()
        return row["order_code"]
    # Tìm qua creator_message_id (form gốc người dùng gửi)
    cur.execute(
        "SELECT order_code FROM orders WHERE chat_id=? AND creator_message_id=?",
        (chat_id, message_id)
    )
    row = cur.fetchone()
    conn.close()
    return row["order_code"] if row else None


def get_order_creator(order_code: str) -> Optional[int]:
    conn = db_connect()
    cur = conn.cursor()
    cur.execute("SELECT creator_user_id FROM orders WHERE order_code=?", (order_code,))
    row = cur.fetchone()
    conn.close()
    return row["creator_user_id"] if row else None


async def cancel_order(bot: Bot, chat_id: int, order_code: str,
                       cancelled_by: str, cancel_msg_id: int) -> None:
    """
    Hủy đơn:
    1. Xóa TẤT CẢ tin nhắn QR của đơn
    2. Xóa tin nhắn "huy" của user
    3. Gửi thông báo "Đã hủy đơn bởi..." — để mãi trong nhóm
    4. Nếu đã gửi collect → edit card thành "Đã hủy"
    5. Cập nhật DB status = cancelled
    """
    conn = db_connect()
    cur = conn.cursor()

    # Lấy tất cả message QR + collect_message_id
    cur.execute(
        "SELECT message_id FROM order_messages WHERE chat_id=? AND order_code=? ORDER BY message_id ASC",
        (chat_id, order_code)
    )
    mids = [r["message_id"] for r in cur.fetchall()]

    cur.execute(
        "SELECT creator_message_id, collect_message_id FROM orders WHERE order_code=?",
        (order_code,)
    )
    crow = cur.fetchone()
    creator_mid     = crow["creator_message_id"] if crow else None
    collect_mid     = crow["collect_message_id"] if crow else None
    conn.close()

    # Xóa tất cả QR
    for mid in mids:
        try:
            await bot.delete_message(chat_id, mid)
        except Exception as _e:
            logger.debug("delete QR %s failed: %s", mid, _e)

    # Xóa form gốc
    if creator_mid and creator_mid != cancel_msg_id:
        try:
            await bot.delete_message(chat_id, creator_mid)
        except Exception:
            pass

    # Xóa tin "huy" của user
    try:
        await bot.delete_message(chat_id, cancel_msg_id)
    except Exception:
        pass

    # Gửi thông báo hủy — để mãi trong Group QR
    cancelled_at = now_local().strftime('%H:%M %d/%m/%Y')
    await bot.send_message(
        chat_id=chat_id,
        text=(
            f"🚫 <b>Đã hủy đơn</b> <code>{order_code}</code>\n"
            f"👤 Hủy bởi: {cancelled_by}\n"
            f"🕐 {cancelled_at}"
        ),
        parse_mode="HTML",
    )

    # Nếu đã gửi card sang Group Collect → edit thành "Đã hủy"
    if collect_mid and COLLECT_GROUP_ID:
        cancelled_text = (
            f"🚫 <b>ĐƠN ĐÃ HỦY</b>\n"
            f"📦 Mã đơn: <b><code>{order_code}</code></b>\n"
            f"👤 Hủy bởi: {cancelled_by}\n"
            f"🕐 {cancelled_at}"
        )
        try:
            await bot.edit_message_text(
                chat_id=COLLECT_GROUP_ID,
                message_id=collect_mid,
                text=cancelled_text,
                parse_mode="HTML",
            )
        except Exception as e:
            logger.warning("Không edit được card collect khi hủy: %s", e)
            # Fallback: reply thông báo vào card
            try:
                await bot.send_message(
                    chat_id=COLLECT_GROUP_ID,
                    text=f"🚫 <b>Đơn <code>{order_code}</code> đã bị hủy</b> bởi {cancelled_by}",
                    parse_mode="HTML",
                    reply_to_message_id=collect_mid,
                )
            except Exception:
                pass

    # Cập nhật DB
    conn = db_connect()
    cur = conn.cursor()
    cur.execute("DELETE FROM order_messages WHERE chat_id=? AND order_code=?", (chat_id, order_code))
    cur.execute("UPDATE orders SET status='cancelled' WHERE order_code=?", (order_code,))
    conn.commit()
    conn.close()


async def delete_order(bot: Bot, chat_id: int, order_code: str) -> None:
    """Xóa hoàn toàn đơn (dùng cho admin xóa cứng nếu cần)."""
    conn = db_connect()
    cur = conn.cursor()
    cur.execute(
        "SELECT message_id FROM order_messages WHERE chat_id=? AND order_code=?",
        (chat_id, order_code)
    )
    mids = [r["message_id"] for r in cur.fetchall()]
    for mid in mids:
        try:
            await bot.delete_message(chat_id, mid)
        except Exception:
            pass
    cur.execute("DELETE FROM order_messages WHERE chat_id=? AND order_code=?", (chat_id, order_code))
    cur.execute("DELETE FROM receivers WHERE order_code=?", (order_code,))
    cur.execute("DELETE FROM orders WHERE order_code=?", (order_code,))
    conn.commit()
    conn.close()


# ─────────────────────────── NOTIFY HỆ THỐNG ────────────────────────────────

async def notify_system(bot: Bot, text: str) -> None:
    """
    Gửi thông báo hệ thống đến NOTIFY_GROUP_ID.
    Dùng cho: khởi động, tắt, kích hoạt nhóm, hủy đơn, lỗi nghiêm trọng.
    """
    if not NOTIFY_GROUP_ID:
        return
    try:
        # Giới hạn độ dài để tránh Telegram reject
        if len(text) > 4000:
            text = text[:3990] + "\n...(truncated)"
        await bot.send_message(NOTIFY_GROUP_ID, text, parse_mode=ParseMode.HTML)
    except Exception as e:
        # Log ra console, không raise để tránh crash main flow
        logger.warning("notify_system thất bại (group %s): %s", NOTIFY_GROUP_ID, e)


# ─────────────────────────── CANCEL LOG ──────────────────────────────────────

def write_cancel_log(entry: dict) -> None:
    """Ghi log hủy đơn vào file JSONL."""
    line = json.dumps(entry, ensure_ascii=False)
    with open(CANCEL_LOG_PATH, "a", encoding="utf-8") as f:
        f.write(line + "\n")


async def notify_cancel(bot: Bot, entry: dict) -> None:
    """Gửi thông báo hủy đơn đến NOTIFY_GROUP_ID."""
    text = (
        "🔴 <b>ĐƠN BỊ HỦY</b>\n"
        f"📦 Mã đơn   : <code>{entry['order_code']}</code>\n"
        f"🏘 Nhóm     : {entry['group_name']}\n"
        f"👤 Hủy bởi  : {entry['cancelled_by_name']} "
        f"(<code>{entry['cancelled_by_id']}</code>)\n"
        f"🕐 Thời gian : {entry['cancelled_at']}\n"
        f"💰 Giá trị  : {format_money(entry['total_amount'])}\n"
        f"📊 Số QR    : {entry['total_qr']}"
    )
    await notify_system(bot, text)


# ─────────────────────────── EXCEL EXPORT ────────────────────────────────────

_HDR_FILL  = PatternFill("solid", fgColor="1F4E79")
_ALT_FILL  = PatternFill("solid", fgColor="E9F0FB")
_SUM_FILL  = PatternFill("solid", fgColor="D6E4F0")
_THIN      = Side(style="thin", color="AAAAAA")
_BORDER    = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_HDR_FONT  = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
_BODY_FONT = Font(name="Calibri", size=10)
_MONEY_FMT = '#,##0'

_DETAIL_HEADERS = [
    "Thời gian", "Nhóm", "Người tạo",
    "NH Người chuyển", "STK Người chuyển", "Tên người chuyển",
    "NH Người nhận", "STK Người nhận", "Tên người nhận",
    "Số tiền", "Số tiền thực", "Nội dung", "Mã Đơn", "Trạng thái",
    "Thời gian HT", "Người xác nhận", "Ghi chú",
]


def _apply_header_row(ws, row_idx: int = 1) -> None:
    for cell in ws[row_idx]:
        cell.font = _HDR_FONT
        cell.fill = _HDR_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _BORDER


def _apply_body_rows(ws, start_row: int = 2) -> None:
    for i, row in enumerate(ws.iter_rows(min_row=start_row), start=0):
        fill = _ALT_FILL if i % 2 == 1 else None
        for cell in row:
            cell.font = _BODY_FONT
            cell.border = _BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=False)
            if fill:
                cell.fill = fill


def export_orders_to_excel(rows: List[sqlite3.Row], output_path: str) -> None:
    wb = Workbook()

    # ── Sheet 1: Chi tiết ─────────────────────────────────────────────────────
    ws_detail = wb.active
    ws_detail.title = "Chi Tiết"
    ws_detail.row_dimensions[1].height = 30
    ws_detail.append(_DETAIL_HEADERS)

    for row in rows:
        ws_detail.append([
            row["created_at"],
            row["group_name"],
            row["creator_name"],
            row["sender_bank"],
            row["sender_account"],
            row["sender_name"],
            row["receiver_bank"],
            row["receiver_account"],
            row["receiver_name"],
            row["amount"],
            row["actual_amount"] or row["amount"],
            row["content"],
            row["order_code"],
            {"cancelled": "Đã hủy", "completed": "Hoàn thành", "sent": "Đã gửi"}.get(row["status"], "Active"),
            row["completed_at"] or "",
            row["completed_by_name"] or "",
            str(row["message_id"] or ""),
        ])

    # Định dạng cột tiền
    for r in range(2, ws_detail.max_row + 1):
        ws_detail.cell(r, 10).number_format = _MONEY_FMT
        ws_detail.cell(r, 11).number_format = _MONEY_FMT

    _apply_header_row(ws_detail)
    _apply_body_rows(ws_detail)
    ws_detail.freeze_panes = "A2"

    col_widths_detail = [20, 22, 18, 14, 20, 24, 14, 20, 24, 14, 14, 18, 16, 12, 18, 18, 12]
    for i, w in enumerate(col_widths_detail, 1):
        ws_detail.column_dimensions[
            ws_detail.cell(1, i).column_letter
        ].width = w

    # ── Sheet 2: Tổng hợp ────────────────────────────────────────────────────
    ws_sum = wb.create_sheet("Tổng Hợp")

    # ── Thu thập dữ liệu ──────────────────────────────────────────────────────
    # Bank nhận: group theo (bank, stk, tên)
    recv_detail: dict[str, dict] = {}   # key = "BANK|STK"
    send_detail: dict[str, dict] = {}   # key = "BANK|STK"
    for row in rows:
        rk = f"{(row['receiver_bank'] or '?').upper()}|{row['receiver_account'] or '?'}"
        sk = f"{(row['sender_bank'] or '?').upper()}|{row['sender_account'] or '?'}"
        if rk not in recv_detail:
            recv_detail[rk] = {
                "bank": (row["receiver_bank"] or "?").upper(),
                "stk": row["receiver_account"] or "?",
                "name": row["receiver_name"] or "?",
                "count": 0, "total": 0,
            }
        recv_detail[rk]["count"] += 1
        recv_detail[rk]["total"] += row["amount"]
        if sk not in send_detail:
            send_detail[sk] = {
                "bank": (row["sender_bank"] or "?").upper(),
                "stk": row["sender_account"] or "?",
                "name": row["sender_name"] or "?",
                "count": 0, "total": 0,
            }
        send_detail[sk]["count"] += 1
        send_detail[sk]["total"] += row["amount"]

    def _write_detail_table(ws, title: str, detail: dict, start_row: int) -> None:
        ws.cell(start_row, 1, title).font = Font(bold=True, size=11, color="1F4E79", name="Calibri")
        hdr = start_row + 1
        for col, val in enumerate(["Ngân hàng", "STK", "Tên TK", "Số lượng QR", "Tổng tiền (đ)"], 1):
            ws.cell(hdr, col, val)
        _apply_header_row(ws, row_idx=hdr)
        data_start = hdr + 1
        for d in sorted(detail.values(), key=lambda x: x["bank"]):
            ws.append([d["bank"], d["stk"], d["name"], d["count"], d["total"]])
        for r in range(data_start, ws.max_row + 1):
            ws.cell(r, 5).number_format = _MONEY_FMT
        _apply_body_rows(ws, start_row=data_start)
        tr = ws.max_row + 1
        ws.cell(tr, 1, "TỔNG CỘNG")
        ws.cell(tr, 4, sum(d["count"] for d in detail.values()))
        ws.cell(tr, 5, sum(d["total"] for d in detail.values()))
        ws.cell(tr, 5).number_format = _MONEY_FMT
        for c in range(1, 6):
            cell = ws.cell(tr, c)
            cell.font = Font(bold=True, name="Calibri", size=10)
            cell.fill = _SUM_FILL
            cell.border = _BORDER

    # ── Ghi 2 bảng ────────────────────────────────────────────────────────────
    _write_detail_table(ws_sum, "THỐNG KÊ THEO NGÂN HÀNG NHẬN", recv_detail, 1)
    ws_sum.append([])
    ws_sum.append([])
    _write_detail_table(ws_sum, "THỐNG KÊ THEO NGÂN HÀNG CHUYỂN", send_detail, ws_sum.max_row + 1)
    ws_sum.append([])

    ws_sum.column_dimensions["A"].width = 16
    ws_sum.column_dimensions["B"].width = 20
    ws_sum.column_dimensions["C"].width = 26
    ws_sum.column_dimensions["D"].width = 13
    ws_sum.column_dimensions["E"].width = 18

    # --- Bảng tổng theo mã đơn ---
    date_row = ws_sum.max_row + 1
    ws_sum.append(["THỐNG KÊ THEO MÃ ĐƠN"])
    ws_sum.cell(date_row, 1).font = Font(bold=True, size=12, color="1F4E79", name="Calibri")
    hdr_row = ws_sum.max_row + 1
    ws_sum.append(["Mã Đơn", "Người tạo", "Số QR", "Tổng tiền (đ)", "Thời gian"])
    _apply_header_row(ws_sum, row_idx=hdr_row)

    order_stats: dict[str, dict] = {}
    for row in rows:
        oc = row["order_code"]
        if oc not in order_stats:
            order_stats[oc] = {
                "creator": row["creator_name"],
                "count": 0, "total": 0,
                "ts": row["created_at"],
            }
        order_stats[oc]["count"] += 1
        order_stats[oc]["total"] += row["amount"]

    ord_start = hdr_row + 1
    for oc, stat in order_stats.items():
        ws_sum.append([oc, stat["creator"], stat["count"], stat["total"], stat["ts"]])

    for r in range(ord_start, ws_sum.max_row + 1):
        ws_sum.cell(r, 4).number_format = _MONEY_FMT

    _apply_body_rows(ws_sum, start_row=ord_start)

    ws_sum.column_dimensions["A"].width = 18
    ws_sum.column_dimensions["B"].width = 22
    ws_sum.column_dimensions["C"].width = 12
    ws_sum.column_dimensions["D"].width = 18
    ws_sum.column_dimensions["E"].width = 20

    ws_sum.freeze_panes = "A3"

    wb.save(output_path)
    logger.info("Excel đã lưu: %s", output_path)


# ─────────────────────────── DB FETCH ────────────────────────────────────────

def fetch_by_code(order_code: str) -> List[sqlite3.Row]:
    conn = db_connect()
    cur = conn.cursor()
    cur.execute("""
        SELECT o.created_at, o.group_name, o.creator_name,
               o.sender_bank, o.sender_account, o.sender_name,
               o.order_code, o.chat_id, o.status,
               o.completed_at, o.completed_by_name,
               r.receiver_bank, r.receiver_account, r.receiver_name,
               r.amount, r.actual_amount, r.content, r.message_id
        FROM orders o
        JOIN receivers r ON o.order_code = r.order_code
        WHERE o.order_code = ?
        ORDER BY r.receiver_index
    """, (order_code,))
    rows = cur.fetchall()
    conn.close()
    return rows


def fetch_by_date(date_str_ddmmyyyy: str, chat_id: Optional[int] = None) -> List[sqlite3.Row]:
    date_obj = datetime.strptime(date_str_ddmmyyyy, "%d/%m/%Y")
    date_key = date_obj.strftime("%Y-%m-%d")
    conn = db_connect()
    cur = conn.cursor()
    if chat_id:
        cur.execute("""
            SELECT o.created_at, o.group_name, o.creator_name,
                   o.sender_bank, o.sender_account, o.sender_name,
                   o.order_code, o.chat_id, o.status,
                   o.completed_at, o.completed_by_name,
                   r.receiver_bank, r.receiver_account, r.receiver_name,
                   r.amount, r.actual_amount, r.content, r.message_id
            FROM orders o
            JOIN receivers r ON o.order_code = r.order_code
            WHERE substr(o.created_at,1,10)=? AND o.chat_id=?
            ORDER BY o.created_at, o.order_code, r.receiver_index
        """, (date_key, chat_id))
    else:
        cur.execute("""
            SELECT o.created_at, o.group_name, o.creator_name,
                   o.sender_bank, o.sender_account, o.sender_name,
                   o.order_code, o.chat_id, o.status,
                   o.completed_at, o.completed_by_name,
                   r.receiver_bank, r.receiver_account, r.receiver_name,
                   r.amount, r.actual_amount, r.content, r.message_id
            FROM orders o
            JOIN receivers r ON o.order_code = r.order_code
            WHERE substr(o.created_at,1,10)=?
            ORDER BY o.created_at, o.order_code, r.receiver_index
        """, (date_key,))
    rows = cur.fetchall()
    conn.close()
    return rows


# ─────────────────────────── DISPATCHER ──────────────────────────────────────


dp = Dispatcher()

@dp.message(Command("start"))
async def cmd_start(message: Message) -> None:
    """Hướng dẫn sử dụng — hoạt động cả DM lẫn group."""
    text = (
        "🤖 <b>QrTC V2</b> — Quản trị hệ thống\n"
        "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"

        "🔱 <b>SUPERADMIN</b>\n"
        "/kichhoat — Kích hoạt bot tại nhóm\n"
        "/tatbot — Tắt bot khỏi nhóm\n"
        "/addadmin [ID] — Thêm global admin\n"
        "/removeadmin [ID] — Xóa global admin\n"
        "/listadmin — Danh sách admin\n"
        "/resetdb — Xóa sạch toàn bộ đơn hàng\n\n"

        "🛡 <b>ADMIN</b>\n"
        "/update — Import danh sách TK whitelist (.txt)\n"
        "/listbank — Xem danh sách TK whitelist\n"
        "/tatbank [STK...] — Tắt 1 hoặc nhiều TK\n"
        "/mobank [STK...] — Mở lại TK\n"
        "/deletebank [STK...] — Xóa hẳn TK khỏi whitelist\n"
        "/fix [index] [tiền] — Sửa số tiền thực nhận\n"
        "/clearold — Dừng nhắc đơn cũ\n"
        "/checkdon [mã] — Xuất Excel theo mã đơn\n"
        "/checkbank [dd/mm/yyyy] — Xuất Excel theo ngày\n\n"

        "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
        "🔄 <b>TRẠNG THÁI ĐƠN</b>\n"
        "⏳ Active → ✅ Đã gửi → 📋 Có bill → ❌ Hủy\n\n"

        "📋 <b>FORMAT FILE WHITELIST (.txt)</b>\n"
        "<code>Bank - MãTB - Tên TK - STK\n"
        "VD: SEABANK - 1436 - Nguyễn Thị A - 000006250473</code>"
    )
    await message.reply(text, parse_mode=ParseMode.HTML)




# ── /kichhoat — Superadmin kích hoạt nhóm ────────────────────────────────────
@dp.message(Command("kichhoat"))
async def cmd_kichhoat(message: Message, bot: Bot) -> None:
    if not message.from_user:
        return
    logger.info("[DEBUG] /kichhoat | user_id=%s | SUPER_ADMIN_IDS=%s | ok=%s",
                message.from_user.id, SUPER_ADMIN_IDS,
                message.from_user.id in SUPER_ADMIN_IDS)
    if message.from_user.id not in SUPER_ADMIN_IDS:
        await message.reply("❌ Chỉ superadmin mới có thể kích hoạt bot.")
        return

    conn = db_connect()
    cur = conn.cursor()
    cur.execute("""
        INSERT OR REPLACE INTO activated_chats (chat_id, activated_by, activated_at)
        VALUES (?,?,?)
    """, (message.chat.id, message.from_user.id, now_local().strftime("%Y-%m-%d %H:%M:%S")))
    conn.commit()
    conn.close()

    g_name = chat_title(message)
    activator = sender_display_name(message)
    await notify_system(
        bot,
        f"✅ <b>Kích hoạt nhóm mới</b>\n"
        f"🏘 Nhóm : {g_name} (<code>{message.chat.id}</code>)\n"
        f"👤 Bởi  : {activator} (<code>{message.from_user.id}</code>)\n"
        f"🕐 Lúc  : {now_local().strftime('%Y-%m-%d %H:%M:%S')}"
    )
    await message.reply("✅ Kích hoạt thành công.", parse_mode=ParseMode.HTML)


# ── /tatbot — Superadmin tắt bot khỏi nhóm ───────────────────────────────────
@dp.message(Command("tatbot"))
async def cmd_tatbot(message: Message, bot: Bot) -> None:
    if not message.from_user:
        return
    if message.from_user.id not in SUPER_ADMIN_IDS:
        await message.reply("❌ Chỉ superadmin mới có thể tắt bot.")
        return

    conn = db_connect()
    cur = conn.cursor()
    cur.execute("DELETE FROM activated_chats WHERE chat_id=?", (message.chat.id,))
    conn.commit()
    conn.close()

    g_name = chat_title(message)
    deactivator = sender_display_name(message)
    await notify_system(
        bot,
        f"⛔ <b>Bot bị tắt tại nhóm</b>\n"
        f"🏘 Nhóm : {g_name} (<code>{message.chat.id}</code>)\n"
        f"👤 Bởi  : {deactivator} (<code>{message.from_user.id}</code>)\n"
        f"🕐 Lúc  : {now_local().strftime('%Y-%m-%d %H:%M:%S')}"
    )
    await message.reply("⛔ Bot đã bị tắt tại nhóm này.")


# ── /addadmin — Thêm admin bot ────────────────────────────────────────────────
@dp.message(Command("addadmin"))
async def cmd_addadmin(message: Message) -> None:
    if not message.from_user:
        return
    # Chỉ superadmin mới /addadmin được
    if message.from_user.id not in SUPER_ADMIN_IDS:
        await message.reply("❌ Chỉ superadmin mới có thể thêm admin bot.")
        return

    target_id = None
    if message.reply_to_message and message.reply_to_message.from_user:
        target_id = message.reply_to_message.from_user.id
    else:
        parts = message.text.split(maxsplit=1)
        if len(parts) == 2 and parts[1].strip().isdigit():
            target_id = int(parts[1].strip())

    if not target_id:
        await message.reply("Dùng: reply vào user + /addadmin  hoặc  /addadmin <user_id>")
        return

    conn = db_connect()
    cur = conn.cursor()
    cur.execute("""
        INSERT OR REPLACE INTO global_admins (user_id, added_by, added_at)
        VALUES (?,?,?)
    """, (target_id, message.from_user.id, now_local().strftime("%Y-%m-%d %H:%M:%S")))
    conn.commit()
    conn.close()

    await message.reply(f"✅ Đã thêm global admin: <code>{target_id}</code>", parse_mode=ParseMode.HTML)
    await notify_system(
        message.bot,
        f"👤 <b>Thêm global admin</b>\n"
        f"ID: <code>{target_id}</code>\n"
        f"Bởi: {sender_display_name(message)}"
    )


# ── /removeadmin — Xóa admin bot ─────────────────────────────────────────────
@dp.message(Command("removeadmin"))
async def cmd_removeadmin(message: Message) -> None:
    if not message.from_user:
        return
    if message.from_user.id not in SUPER_ADMIN_IDS:
        await message.reply("❌ Chỉ superadmin mới có thể xóa admin bot.")
        return

    target_id = None
    if message.reply_to_message and message.reply_to_message.from_user:
        target_id = message.reply_to_message.from_user.id
    else:
        parts = message.text.split(maxsplit=1)
        if len(parts) == 2 and parts[1].strip().isdigit():
            target_id = int(parts[1].strip())

    if not target_id:
        await message.reply("Dùng: reply vào user + /removeadmin  hoặc  /removeadmin <user_id>")
        return

    if target_id in SUPER_ADMIN_IDS:
        await message.reply("❌ Không thể xóa superadmin.")
        return

    conn = db_connect()
    cur = conn.cursor()
    cur.execute("DELETE FROM global_admins WHERE user_id=?", (target_id,))
    conn.commit()
    conn.close()
    await message.reply(f"✅ Đã xóa global admin: <code>{target_id}</code>", parse_mode=ParseMode.HTML)
    await notify_system(
        message.bot,
        f"🗑 <b>Xóa global admin</b>\n"
        f"ID: <code>{target_id}</code>\n"
        f"Bởi: {sender_display_name(message)}"
    )


# ── /listadmin — Danh sách admin ─────────────────────────────────────────────
@dp.message(Command("listadmin"))
async def cmd_listadmin(message: Message) -> None:
    conn = db_connect()
    cur = conn.cursor()
    cur.execute("SELECT user_id, added_by, added_at, note FROM global_admins ORDER BY added_at")
    rows = cur.fetchall()
    conn.close()

    lines = ["<b>👑 Danh sách admin bot (toàn hệ thống)</b>\n"]
    if SUPER_ADMIN_IDS:
        lines.append("🔱 <b>Superadmin:</b>")
        for uid in sorted(SUPER_ADMIN_IDS):
            lines.append(f"  • <code>{uid}</code>")
        lines.append("")

    if rows:
        lines.append("🛡 <b>Global Admin:</b>")
        for r in rows:
            note = f" — {r['note']}" if r["note"] else ""
            lines.append(
                f"  • <code>{r['user_id']}</code>"
                f"  — thêm bởi <code>{r['added_by']}</code>"
                f"  lúc {r['added_at']}{note}"
            )
    else:
        lines.append("Chưa có global admin nào.")

    await message.reply("\n".join(lines), parse_mode=ParseMode.HTML)


# ── /checkdon ─────────────────────────────────────────────────────────────────
@dp.message(Command("checkdon"))
async def cmd_checkdon(message: Message) -> None:
    if not chat_is_active(message.chat.id):
        return

    parts = message.text.split(maxsplit=1)
    if len(parts) < 2:
        await message.reply("Dùng: /checkdon MÃ_ĐƠN\nVí dụ: /checkdon XT01120526")
        return

    code = parts[1].strip().upper()
    rows = fetch_by_code(code)
    if not rows:
        await message.reply(f"Không tìm thấy mã đơn <code>{code}</code>.", parse_mode=ParseMode.HTML)
        return

    with tempfile.TemporaryDirectory() as tmpdir:
        fpath = os.path.join(tmpdir, f"checkdon_{code}.xlsx")
        export_orders_to_excel(rows, fpath)
        await message.reply_document(
            FSInputFile(fpath, filename=f"checkdon_{code}.xlsx"),
            caption=f"📄 Chi tiết mã đơn <code>{code}</code>",
            parse_mode=ParseMode.HTML,
        )


# ── /checkbank ────────────────────────────────────────────────────────────────
@dp.message(Command("checkbank"))
async def cmd_checkbank(message: Message) -> None:
    if not chat_is_active(message.chat.id):
        return

    parts = message.text.split(maxsplit=1)
    date_str = now_local().strftime("%d/%m/%Y") if len(parts) == 1 else parts[1].strip()

    if not re.fullmatch(r"\d{2}/\d{2}/\d{4}", date_str):
        await message.reply("Sai định dạng ngày.\nDùng: /checkbank dd/mm/yyyy")
        return

    try:
        rows = fetch_by_date(date_str, chat_id=message.chat.id)
    except ValueError:
        await message.reply("Ngày không hợp lệ. Dùng: /checkbank dd/mm/yyyy")
        return

    if not rows:
        await message.reply(f"Không có dữ liệu ngày <b>{date_str}</b>.", parse_mode=ParseMode.HTML)
        return

    safe_date = date_str.replace("/", "-")
    with tempfile.TemporaryDirectory() as tmpdir:
        fpath = os.path.join(tmpdir, f"checkbank_{safe_date}.xlsx")
        export_orders_to_excel(rows, fpath)
        await message.reply_document(
            FSInputFile(fpath, filename=f"checkbank_{safe_date}.xlsx"),
            caption=f"📄 Checkbank ngày <b>{date_str}</b>",
            parse_mode=ParseMode.HTML,
        )


# ── Handler chính: form tạo QR + hủy đơn ─────────────────────────────────────




@dp.message(Command("fix"))
async def cmd_fix_amount(message: Message, bot: Bot) -> None:
    """
    Sửa số tiền thực nhận của 1 người nhận trong đơn.
    Dùng trong Group Collect, reply vào card đơn.
    Chỉ admin/superadmin được dùng.
    Format: /fix <index> <số tiền>
    Ví dụ: /fix 2 101,500,000
    """
    if message.chat.id != COLLECT_GROUP_ID:
        return
    if not message.from_user:
        return

    # Kiểm tra quyền: superadmin hoặc bot_admin được add bằng /addadmin
    uid = message.from_user.id
    if not await is_bot_admin(COLLECT_GROUP_ID, uid):
        await message.reply("❌ Chỉ admin bot mới được dùng lệnh /fix.")
        return

    if not message.reply_to_message:
        await message.reply("❌ Hãy reply vào card đơn rồi dùng /fix <số thứ tự> <số tiền thực>")
        return

    # Parse: /fix 1 101,500,000
    #        /fix 1,2,3 101,500,000     → cùng số tiền cho nhiều TK
    #        /fix 1:101500000 2:95000000 → từng TK số tiền khác nhau
    text_body = message.text.split(maxsplit=1)[1].strip() if len(message.text.split(maxsplit=1)) > 1 else ""
    if not text_body:
        await message.reply(
            "❌ Cú pháp /fix:\n• /fix 1 101,500,000\n• /fix 1,2,3 101,500,000  (cùng tiền)\n• /fix 1:101500000 2:95000000  (tiền khác nhau)"
        )
        return

    replied_id = message.reply_to_message.message_id

    # Tìm đơn theo collect_message_id
    conn = db_connect()
    cur = conn.cursor()
    cur.execute("SELECT order_code FROM orders WHERE collect_message_id=?", (replied_id,))
    row = cur.fetchone()
    if not row:
        conn.close()
        await message.reply("❌ Không tìm thấy đơn tương ứng với card này.")
        return
    order_code = row["order_code"]

    # Parse cú pháp fix_items: list of (index, amount)
    fix_items: list[tuple[int, int]] = []
    parse_errors: list[str] = []

    if ":" in text_body:
        # Dạng: 1:101500000 2:95000000
        for token in text_body.split():
            if ":" not in token:
                continue
            idx_str, amt_str = token.split(":", 1)
            try:
                idx = int(idx_str.strip())
                amt = parse_amount(amt_str.strip())
                if amt is None:
                    parse_errors.append(f"Số tiền '{amt_str}' không hợp lệ")
                else:
                    fix_items.append((idx, amt))
            except ValueError:
                parse_errors.append(f"Index '{idx_str}' không hợp lệ")
    else:
        # Dạng: 1,2,3 101,500,000  hoặc  1 101,500,000
        tokens = text_body.split()
        # Tìm index block (phần đầu không phải số tiền)
        idx_part = tokens[0]
        amt_part = " ".join(tokens[1:]) if len(tokens) > 1 else ""
        amt = parse_amount(amt_part)
        if amt is None:
            await message.reply(
                "❌ Không nhận diện được số tiền.\nVí dụ: /fix 1,2,3 101,500,000"
            )
            return
        for idx_str in re.split(r"[,;\s]+", idx_part):
            if not idx_str.strip():
                continue
            try:
                fix_items.append((int(idx_str.strip()), amt))
            except ValueError:
                parse_errors.append(f"Index '{idx_str}' không hợp lệ")

    if not fix_items:
        await message.reply("❌ Không có TK nào được fix. Kiểm tra lại cú pháp.")
        return

    # Cập nhật từng receiver
    results: list[str] = []
    conn = db_connect()
    cur = conn.cursor()
    for recv_index, actual in fix_items:
        cur.execute(
            "SELECT amount FROM receivers WHERE order_code=? AND receiver_index=?",
            (order_code, recv_index)
        )
        recv_row = cur.fetchone()
        if not recv_row:
            results.append(f"❌ TK {recv_index}: không tìm thấy trong đơn {order_code}")
            continue
        old_amount = recv_row["amount"]
        cur.execute(
            "UPDATE receivers SET actual_amount=? WHERE order_code=? AND receiver_index=?",
            (actual, order_code, recv_index)
        )
        results.append(f"✅ TK {recv_index}: {format_money(old_amount)} → {format_money(actual)}")
    conn.commit()
    conn.close()

    # Biến old_amount để dùng cho log (lấy từ lần cuối, nếu 1 TK)
    old_amount = fix_items[0][1] if fix_items else 0  # placeholder cho log

    # Rebuild và edit card
    # Lấy trạng thái hiện tại
    conn = db_connect()
    cur = conn.cursor()
    cur.execute("SELECT status, completed_by_name, completed_at FROM orders WHERE order_code=?", (order_code,))
    o = cur.fetchone()
    conn.close()

    status = o["status"] if o else "sent"
    completer = o["completed_by_name"] or "" if o else ""
    completed_at_str = ""
    if o and o["completed_at"]:
        try:
            from datetime import datetime as _dt
            completed_at_str = _dt.strptime(o["completed_at"], "%Y-%m-%d %H:%M:%S").strftime("%H:%M %d/%m/%Y")
        except Exception:
            completed_at_str = o["completed_at"]

    new_text = _build_collect_html(
        order_code,
        status=status,
        completer=completer,
        completed_at=completed_at_str,
    )

    try:
        await bot.edit_message_text(
            chat_id=COLLECT_GROUP_ID,
            message_id=replied_id,
            text=new_text,
            parse_mode="HTML",
        )
    except Exception as e:
        logger.warning("Không edit được card sau /fix: %s", e)

    fixer = message.from_user.username or message.from_user.full_name or str(uid)
    result_text = "\n".join(results)
    if parse_errors:
        result_text += "\n⚠️ " + "\n⚠️ ".join(parse_errors)
    await message.reply(
        f"📝 <b>Fix đơn <code>{order_code}</code></b>\n"
        f"{result_text}\n"
        f"✏️ Sửa bởi: @{fixer if not fixer.startswith('@') else fixer[1:]}",
        parse_mode=ParseMode.HTML,
    )
    logger.info("/fix đơn %s bởi %s: %s", order_code, fixer, results)

@dp.message(F.document)
async def handle_whitelist_file(message: Message, bot: Bot) -> None:
    """Nhận file txt whitelist — hoạt động cả DM lẫn group kích hoạt."""
    if not message.from_user or not message.document:
        return

    uid = message.from_user.id
    now_ts = time.monotonic()

    logger.info("[UPDATE] document từ uid=%s | session=%s | file=%s",
                uid, uid in _update_sessions,
                message.document.file_name or "?")

    # Chỉ xử lý nếu user đang trong session /update
    if uid not in _update_sessions:
        return
    if now_ts - _update_sessions[uid] > _UPDATE_TIMEOUT:
        del _update_sessions[uid]
        await message.reply("⏰ Session đã hết hạn. Gõ /update lại nhé.")
        return

    # Chỉ nhận file .txt
    fname = message.document.file_name or ""
    if not fname.lower().endswith(".txt"):
        await message.reply("⚠️ Chỉ nhận file <code>.txt</code> thôi nhé 🩷",
                            parse_mode=ParseMode.HTML)
        return

    # Download file
    import io
    file_info = await bot.get_file(message.document.file_id)
    file_bytes = await bot.download_file(file_info.file_path)
    content = file_bytes.read().decode("utf-8", errors="ignore")

    # Parse từng dòng: bank - mã thiết bị - tên TK - STK
    # VD: shinhanbank -1447 - Dương Văn Dũng CO - 700040310032
    valid = []
    errors = []
    for i, line in enumerate(content.splitlines(), 1):
        line = line.strip()
        if not line or line.startswith("#"):
            continue
        parts = [p.strip() for p in re.split(r"\s*-\s*", line) if p.strip()]
        if len(parts) < 2:
            errors.append(f"Dòng {i}: '{line[:50]}' — không đủ thông tin")
            continue

        # Format: bank - device_code - name - account
        # VD: shinhanbank - 1447 - Dương Văn Dũng CO - 700040310032
        device_code = ""
        if len(parts) >= 4:
            bank        = parts[0]
            device_code = parts[1]            # mã thiết bị
            name        = " ".join(parts[2:-1])
            account     = parts[-1]
        elif len(parts) == 3:
            bank, name, account = parts[0], parts[1], parts[2]
        elif len(parts) == 2:
            bank, name, account = parts[0], "", parts[1]
        else:
            errors.append(f"Dòng {i}: '{line[:50]}' — không đọc được")
            continue

        account_clean = re.sub(r"\D", "", account)
        if not account_clean or len(account_clean) < 6:
            errors.append(f"Dòng {i}: STK '{account}' không hợp lệ (cần >= 6 chữ số)")
            continue
        valid.append({"bank": bank, "device_code": device_code, "account": account_clean, "name": name})

    if not valid:
        await message.reply(
            f"❌ Không đọc được TK nào từ file.\n"
            f"Lỗi:\n" + "\n".join(errors[:10]),
            parse_mode=ParseMode.HTML
        )
        return

    # Preview + confirm button
    from aiogram.types import InlineKeyboardMarkup, InlineKeyboardButton
    import json as _json

    preview_lines = [
        f"📋 <b>Preview import whitelist</b>\n",
        f"✅ Hợp lệ: <b>{len(valid)}</b> TK",
        f"⚠️ Lỗi: <b>{len(errors)}</b> dòng\n",
    ]
    for r in valid[:15]:
        preview_lines.append(f"  • {r['bank'].upper()} — <code>{r['account']}</code> — {r['name']}")
    if len(valid) > 15:
        preview_lines.append(f"  ... và {len(valid) - 15} TK nữa")
    if errors:
        preview_lines.append("\n⚠️ Một số dòng lỗi:")
        for e in errors[:5]:
            preview_lines.append(f"  • {e}")

    # Lưu tạm data vào session (dùng message_id làm key)
    session_key = f"wl_{uid}_{message.message_id}"
    # Lưu vào DB tạm (dùng bảng đơn giản)
    conn = db_connect()
    cur = conn.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS import_sessions (
            key TEXT PRIMARY KEY, data TEXT, created_at TEXT
        )""")
    cur.execute("INSERT OR REPLACE INTO import_sessions VALUES (?,?,?)",
        (session_key, _json.dumps(valid, ensure_ascii=False),
         now_local().strftime("%Y-%m-%d %H:%M:%S")))
    conn.commit()
    conn.close()

    kb = InlineKeyboardMarkup(inline_keyboard=[[
        InlineKeyboardButton(text="✅ Xác nhận import", callback_data=f"wl_import:{session_key}"),
        InlineKeyboardButton(text="❌ Huỷ",             callback_data=f"wl_cancel:{session_key}"),
    ]])
    await message.reply(
        "\n".join(preview_lines),
        parse_mode=ParseMode.HTML,
        reply_markup=kb
    )


@dp.callback_query(lambda c: c.data and c.data.startswith("wl_import:"))
async def cb_wl_import(callback: CallbackQuery, bot: Bot) -> None:
    """Xác nhận import whitelist từ file txt."""
    import json as _json
    session_key = callback.data.split(":", 1)[1]

    conn = db_connect()
    cur = conn.cursor()
    cur.execute("SELECT data FROM import_sessions WHERE key=?", (session_key,))
    row = cur.fetchone()
    if not row:
        await callback.answer("Session đã hết hạn.", show_alert=True)
        conn.close()
        return

    records = _json.loads(row["data"])
    ts = now_local().strftime("%Y-%m-%d %H:%M:%S")

    # Import: INSERT OR REPLACE (cập nhật nếu đã có)
    for r in records:
        cur.execute("""
            INSERT INTO receiver_whitelist (bank, device_code, account, name, is_active, added_at, added_by)
            VALUES (?,?,?,?,1,?,?)
            ON CONFLICT(account) DO UPDATE SET
                bank=excluded.bank, device_code=excluded.device_code,
                name=excluded.name, is_active=1, added_at=excluded.added_at
        """, (r["bank"], r.get("device_code",""), r["account"], r["name"], ts, callback.from_user.id))

    cur.execute("DELETE FROM import_sessions WHERE key=?", (session_key,))
    conn.commit()
    conn.close()

    # Xoá session
    uid = callback.from_user.id
    _update_sessions.pop(uid, None)

    importer = callback.from_user.username or str(callback.from_user.id)
    await callback.message.edit_reply_markup(reply_markup=None)
    await callback.message.reply(
        f"✅ <b>Đã import {len(records)} TK vào whitelist!</b>\n"
        f"👤 Bởi: @{importer}  •  {ts}",
        parse_mode=ParseMode.HTML
    )
    await notify_system(bot,
        f"📂 <b>Import whitelist</b>\n"
        f"✅ {len(records)} TK được import\n"
        f"👤 @{importer}  •  {ts}\n"
        f"💬 Từ: {callback.message.chat.title or 'DM'}"
    )
    await callback.answer("✅ Import thành công!")


@dp.callback_query(lambda c: c.data and c.data.startswith("wl_cancel:"))
async def cb_wl_cancel(callback: CallbackQuery) -> None:
    session_key = callback.data.split(":", 1)[1]
    conn = db_connect()
    cur = conn.cursor()
    cur.execute("DELETE FROM import_sessions WHERE key=?", (session_key,))
    conn.commit()
    conn.close()
    _update_sessions.pop(callback.from_user.id, None)
    await callback.message.edit_reply_markup(reply_markup=None)
    await callback.answer("❌ Đã huỷ import.")

@dp.message(F.photo)
async def handle_bill_photo(message: Message, bot: Bot) -> None:
    """
    Detect nhân viên reply ảnh vào card đơn trong Group Collect.
    Bất kỳ ai trong group collect đều có thể xác nhận.
    """
    if not message.from_user:
        return
    if message.chat.id != COLLECT_GROUP_ID:
        return
    if not message.reply_to_message:
        return

    replied_id = message.reply_to_message.message_id

    # Tìm đơn theo collect_message_id
    conn = db_connect()
    cur = conn.cursor()
    cur.execute(
        "SELECT order_code, chat_id, qr_message_id, button_message_id, status FROM orders WHERE collect_message_id=?",
        (replied_id,)
    )
    row = cur.fetchone()
    conn.close()

    if not row or row["status"] in ("completed", "cancelled"):
        return

    order_code  = row["order_code"]
    qr_chat_id  = row["chat_id"]
    qr_msg_id   = row["qr_message_id"]
    btn_msg_id  = row["button_message_id"]
    completer  = message.from_user.username or message.from_user.full_name or str(message.from_user.id)
    completed_at = now_local().strftime("%H:%M %d/%m/%Y")

    # Cập nhật DB
    conn = db_connect()
    cur = conn.cursor()
    cur.execute(
        "UPDATE orders SET status='completed', completed_at=?, completed_by_name=? WHERE order_code=?",
        (now_local().strftime("%Y-%m-%d %H:%M:%S"), completer, order_code)
    )
    conn.commit()
    conn.close()

    # Edit card trong Group Collect — dùng HTML nhất quán với lúc gửi
    new_text = _build_collect_html(
        order_code, status="completed",
        completer=completer, completed_at=completed_at
    )
    try:
        await bot.edit_message_text(
            chat_id=COLLECT_GROUP_ID,
            message_id=replied_id,
            text=new_text,
            parse_mode="HTML",
        )
    except Exception as edit_err:
        logger.warning("Không edit được card collect: %s", edit_err)
        try:
            await bot.send_message(
                chat_id=COLLECT_GROUP_ID,
                text=f"✅ <b>Đã có bill</b> — <code>{order_code}</code> — {completer} — {completed_at}",
                parse_mode="HTML",
                reply_to_message_id=replied_id,
            )
        except Exception:
            pass

    # Update button trong Group QR (dùng button_message_id, không phải qr_message_id)
    target_mid = btn_msg_id or qr_msg_id
    if target_mid and qr_chat_id:
        from aiogram.types import InlineKeyboardMarkup, InlineKeyboardButton
        kb = InlineKeyboardMarkup(inline_keyboard=[[
            InlineKeyboardButton(
                text=f"✅ Đã có bill — {completed_at}",
                callback_data=f"noop:{order_code}"
            ),
        ]])
        try:
            await bot.edit_message_reply_markup(
                chat_id=qr_chat_id,
                message_id=target_mid,
                reply_markup=kb,
            )
        except Exception as e:
            logger.warning("Không update được button Group QR: %s", e)

    logger.info("Đơn %s hoàn thành bởi %s", order_code, completer)


# ─────────────────────────── WHITELIST BANK HANDLERS ────────────────────────

# Track session /update: user_id đang chờ gửi file
_update_sessions: dict[int, float] = {}   # user_id → timestamp bắt đầu chờ
_UPDATE_TIMEOUT = 300   # 5 phút


@dp.message(Command("listbank"))
async def cmd_listbank(message: Message) -> None:
    if not chat_is_active(message.chat.id):
        return


    conn = db_connect()
    cur = conn.cursor()
    cur.execute("""
        SELECT bank, device_code, account, name, is_active
        FROM receiver_whitelist
        ORDER BY added_at ASC
    """)
    rows = cur.fetchall()
    conn.close()

    if not rows:
        await message.reply("📋 Whitelist trống. Dùng /update để import.", parse_mode=ParseMode.HTML)
        return

    n_open  = sum(1 for r in rows if r["is_active"])
    n_close = len(rows) - n_open

    lines = [
        f"📋 <b>Whitelist TK người nhận</b>  ✅ {n_open} Mở  |  ❌ {n_close} Tắt",
        "─────────────────────────────",
    ]
    for i, r in enumerate(rows, 1):
        status  = "🟢" if r["is_active"] else "🔴"
        bank    = (r["bank"] or "").upper()
        dev     = r["device_code"] or ""
        dev_str = f" <code>{dev}</code>" if dev else ""
        lines.append(
            f"{i}. {status} <b>[{bank}]</b>{dev_str}  <b>{r['name']}</b>  —  <code>{r['account']}</code>"
        )

    # Gửi theo chunk nếu quá dài
    chunks, current = [], ""
    for line in lines:
        if len(current) + len(line) + 1 > 3800:
            chunks.append(current)
            current = line
        else:
            current = (current + "\n" + line) if current else line
    if current:
        chunks.append(current)

    for i, chunk in enumerate(chunks):
        suffix = f"\n<i>Trang {i+1}/{len(chunks)}</i>" if len(chunks) > 1 else ""
        await message.reply(chunk + suffix, parse_mode=ParseMode.HTML)


@dp.message(Command("deletebank"))
async def cmd_deletebank(message: Message) -> None:
    """Xóa hẳn 1 hoặc nhiều STK khỏi whitelist. /deletebank STK1 STK2"""
    if not message.from_user:
        return
    is_dm = message.chat.id > 0
    if not is_dm and not chat_is_active(message.chat.id):
        return
    if not await is_bot_admin(0, message.from_user.id):
        await message.reply("❌ Chỉ admin mới dùng được lệnh này.")
        return

    parts = message.text.split()[1:]
    if not parts:
        await message.reply("Dùng: /deletebank <STK1> <STK2> ...\nVí dụ: /deletebank 9876543210")
        return

    conn = db_connect()
    cur = conn.cursor()
    results = []
    for stk in parts:
        stk = stk.strip()
        cur.execute("SELECT name, bank FROM receiver_whitelist WHERE account=?", (stk,))
        row = cur.fetchone()
        if not row:
            results.append(f"❌ <code>{stk}</code> không có trong whitelist")
        else:
            cur.execute("DELETE FROM receiver_whitelist WHERE account=?", (stk,))
            results.append(f"🗑 <code>{stk}</code> — {row['name']} ({(row['bank'] or '').upper()}) đã xóa")
    conn.commit()
    conn.close()

    result_text = "\n".join(results)
    await message.reply(f"<b>Kết quả /deletebank:</b>\n{result_text}", parse_mode=ParseMode.HTML)
    await notify_system(message.bot,
        f"🗑 <b>/deletebank</b> bởi {sender_display_name(message)}\n{result_text}")



@dp.message(Command("tatbank"))
async def cmd_tatbank(message: Message) -> None:
    """Tắt 1 hoặc nhiều STK. /tatbank STK1 STK2 STK3"""
    if not message.from_user: return
    is_dm = message.chat.id > 0
    if not is_dm and not chat_is_active(message.chat.id):
        return
    if not await is_bot_admin(0, message.from_user.id):
        await message.reply("❌ Chỉ admin mới dùng được lệnh này.")
        return

    parts = message.text.split()[1:]
    if not parts:
        await message.reply("Dùng: /tatbank <STK1> <STK2> ...\nVí dụ: /tatbank 9876543210 1122334455")
        return

    conn = db_connect()
    cur = conn.cursor()
    results = []
    for stk in parts:
        stk = stk.strip()
        cur.execute("SELECT name FROM receiver_whitelist WHERE account=?", (stk,))
        row = cur.fetchone()
        if not row:
            results.append(f"❌ <code>{stk}</code> không có trong whitelist")
        else:
            cur.execute("UPDATE receiver_whitelist SET is_active=0 WHERE account=?", (stk,))
            results.append(f"🔴 <code>{stk}</code> — {row['name']} đã tắt")
    conn.commit()
    conn.close()

    result_text = "\n".join(results)
    await message.reply(f"<b>Kết quả /tatbank:</b>\n{result_text}", parse_mode=ParseMode.HTML)
    await notify_system(message.bot,
        f"🔴 <b>/tatbank</b> bởi {sender_display_name(message)}\n{result_text}")


@dp.message(Command("mobank"))
async def cmd_mobank(message: Message) -> None:
    """Mở lại 1 hoặc nhiều STK. /mobank STK1 STK2"""
    if not message.from_user: return
    is_dm = message.chat.id > 0
    if not is_dm and not chat_is_active(message.chat.id):
        return
    if not await is_bot_admin(0, message.from_user.id):
        await message.reply("❌ Chỉ admin mới dùng được lệnh này.")
        return

    parts = message.text.split()[1:]
    if not parts:
        await message.reply("Dùng: /mobank <STK1> <STK2> ...\nVí dụ: /mobank 9876543210")
        return

    conn = db_connect()
    cur = conn.cursor()
    results = []
    for stk in parts:
        stk = stk.strip()
        cur.execute("SELECT name FROM receiver_whitelist WHERE account=?", (stk,))
        row = cur.fetchone()
        if not row:
            results.append(f"❌ <code>{stk}</code> không có trong whitelist")
        else:
            cur.execute("UPDATE receiver_whitelist SET is_active=1 WHERE account=?", (stk,))
            results.append(f"🟢 <code>{stk}</code> — {row['name']} đã mở")
    conn.commit()
    conn.close()

    result_text = "\n".join(results)
    await message.reply(f"<b>Kết quả /mobank:</b>\n{result_text}", parse_mode=ParseMode.HTML)
    await notify_system(message.bot,
        f"🟢 <b>/mobank</b> bởi {sender_display_name(message)}\n{result_text}")


@dp.message(Command("update"))
async def cmd_update(message: Message) -> None:
    """
    Import whitelist TK.
    - DM với bot: chỉ admin/superadmin
    - Group: phải kích hoạt + admin/superadmin
    """
    if not message.from_user:
        return
    uid = message.from_user.id
    is_dm = message.chat.id > 0
    is_active = chat_is_active(message.chat.id)
    is_admin = await is_bot_admin(0, uid)
    logger.info("[UPDATE] uid=%s is_dm=%s is_active=%s is_admin=%s SUPER=%s",
                uid, is_dm, is_active, is_admin, uid in SUPER_ADMIN_IDS)
    # Cho phép DM hoặc group đã kích hoạt
    if not is_dm and not is_active:
        await message.reply("❌ Nhóm chưa được kích hoạt. Superadmin dùng /kichhoat trước.")
        return
    if not is_admin:
        await message.reply("❌ Chỉ admin mới dùng được lệnh này.")
        return

    uid = message.from_user.id
    now_ts = time.monotonic()

    # Kiểm tra có session đang chờ không
    for existing_uid, ts in list(_update_sessions.items()):
        if now_ts - ts < _UPDATE_TIMEOUT:
            if existing_uid != uid:
                await message.reply(
                    f"⏳ Đang chờ file từ user <code>{existing_uid}</code>. "
                    f"Vui lòng đợi hoặc chờ {int((_UPDATE_TIMEOUT - (now_ts - ts)) / 60) + 1} phút nữa.",
                    parse_mode=ParseMode.HTML
                )
                return
        else:
            del _update_sessions[existing_uid]

    _update_sessions[uid] = now_ts
    await message.reply(
        "📂 <b>Import whitelist TK</b>\n\n"
        "Bạn iu hãy gửi file <code>.txt</code> danh sách TK cho tui nhé 🩷\n\n"
        "<b>Format mỗi dòng:</b>\n"
        "<code>VCB - 9876543210 - Nguyễn Văn A</code>\n"
        "<code>TCB - 1122334455 - Trần Thị B</code>\n\n"
        "⏰ Session hết hạn sau 5 phút.",
        parse_mode=ParseMode.HTML
    )





@dp.message(Command("clearold"))
async def cmd_clearold(message: Message) -> None:
    """Expire tất cả đơn active/sent từ ngày trước. Dùng trong group kích hoạt hoặc DM."""
    if not message.from_user:
        return
    is_dm = message.chat.id > 0
    if not is_dm and not chat_is_active(message.chat.id):
        return
    if not await is_bot_admin(0, message.from_user.id):
        await message.reply("❌ Chỉ admin mới dùng được lệnh này.")
        return

    conn = db_connect()
    cur = conn.cursor()
    cur.execute("""
        UPDATE orders SET status = 'expired'
        WHERE status IN ('active', 'sent')
          AND date(created_at) < date('now', 'localtime')
    """)
    count = cur.rowcount
    conn.commit()
    conn.close()

    await message.reply(
        f"🧹 Đã dừng nhắc nhở <b>{count}</b> đơn cũ.\n"
        f"Các thông báo sẽ không còn nhảy nữa.",
        parse_mode=ParseMode.HTML
    )


@dp.message(Command("resetdb"))
async def cmd_resetdb(message: Message) -> None:
    """Xóa sạch TẤT CẢ đơn cũ khỏi DB. Chỉ superadmin."""
    if not message.from_user:
        return
    if not is_superadmin(message.from_user.id):
        await message.reply("❌ Chỉ superadmin mới dùng được lệnh này.")
        return
    is_dm = message.chat.id > 0
    if not is_dm and not chat_is_active(message.chat.id):
        return

    from aiogram.types import InlineKeyboardMarkup, InlineKeyboardButton
    kb = InlineKeyboardMarkup(inline_keyboard=[[
        InlineKeyboardButton(text="✅ Xác nhận xóa hết", callback_data="resetdb:confirm"),
        InlineKeyboardButton(text="❌ Huỷ",              callback_data="resetdb:cancel"),
    ]])
    await message.reply(
        "⚠️ <b>Xác nhận xóa toàn bộ dữ liệu đơn hàng?</b>\n"
        "Hành động này không thể hoàn tác.",
        parse_mode=ParseMode.HTML,
        reply_markup=kb
    )


@dp.callback_query(lambda c: c.data and c.data.startswith("resetdb:"))
async def cb_resetdb(callback: CallbackQuery, bot: Bot) -> None:
    if not is_superadmin(callback.from_user.id):
        await callback.answer("❌ Không có quyền.", show_alert=True)
        return
    action = callback.data.split(":")[1]
    if action == "cancel":
        await callback.message.edit_reply_markup(reply_markup=None)
        await callback.answer("❌ Đã huỷ.")
        return

    conn = db_connect()
    cur = conn.cursor()
    cur.execute("SELECT COUNT(*) as cnt FROM orders")
    total = cur.fetchone()["cnt"]
    cur.execute("DELETE FROM order_messages")
    cur.execute("DELETE FROM receivers")
    cur.execute("DELETE FROM orders")
    conn.commit()
    conn.close()

    await callback.message.edit_reply_markup(reply_markup=None)
    await callback.message.reply(
        f"🗑 <b>Đã xóa {total} đơn hàng khỏi DB.</b>\n"
        f"Bot sẽ không còn nhắc nhở đơn cũ nữa.",
        parse_mode=ParseMode.HTML
    )
    await notify_system(bot,
        f"🗑 <b>resetdb</b> bởi {callback.from_user.username or callback.from_user.id}\n"
        f"Đã xóa {total} đơn"
    )
    await callback.answer("✅ Đã xóa sạch!")


def check_whitelist(parsed: Dict[str, Any]) -> list:
    """
    Kiểm tra STK người nhận theo WHITELIST_MODE.

    Case 1/3: check từng TK trong parsed["receivers"]
    Case 2:   check TK người nhận duy nhất = parsed["sender"]
              (vì parsed["sender"] là TK nhận tiền trong Case 2)

    Modes:
      blacklist (mặc định): chặn STK có is_active=0
      strict:               chỉ cho phép STK có is_active=1
    """
    conn = db_connect()
    cur = conn.cursor()
    errors = []

    case = parsed.get("case", 1)
    # Case 2: người nhận duy nhất là parsed["sender"]
    accounts_to_check = (
        [parsed["sender"]] if case == 2 else parsed["receivers"]
    )

    for r in accounts_to_check:
        cur.execute(
            "SELECT is_active FROM receiver_whitelist WHERE account=?",
            (r["account"],)
        )
        row = cur.fetchone()

        if WHITELIST_MODE == "strict":
            if not row:
                errors.append(
                    f"⚠️ STK <code>{r['account']}</code> (<b>{r['name']}</b>) "
                    f"không có trong danh sách được phép."
                )
            elif row["is_active"] == 0:
                errors.append(
                    f"🔴 STK <code>{r['account']}</code> (<b>{r['name']}</b>) "
                    f"đang bị tắt, không thể tạo QR."
                )
        else:
            if row and row["is_active"] == 0:
                errors.append(
                    f"🔴 STK <code>{r['account']}</code> (<b>{r['name']}</b>) "
                    f"đang bị tắt. Admin dùng /mobank để mở lại."
                )

    conn.close()
    return errors

@dp.message(F.text)
async def handle_text(message: Message, bot: Bot) -> None:
    if not message.from_user or not message.text:
        return

    text = message.text.strip()
    # DEBUG — xoá sau khi xác nhận hoạt động
    logger.info(
        "[DEBUG] Nhận message | chat_id=%s | user_id=%s | active=%s | text=%r",
        message.chat.id,
        message.from_user.id,
        chat_is_active(message.chat.id),
        text[:80],
    )

    # --- Hủy đơn (reply "huy" / "hủy") ---
    if message.reply_to_message and text.lower() in {"huy", "hủy"}:
        if not chat_is_active(message.chat.id):
            return

        replied_id = message.reply_to_message.message_id
        order_code = await get_order_by_message(message.chat.id, replied_id)
        if not order_code:
            await message.reply("Không tìm thấy đơn để hủy.")
            return

        creator_id = get_order_creator(order_code)
        uid = message.from_user.id
        allowed = (uid == creator_id) or await is_bot_admin(message.chat.id, uid)

        if not allowed:
            warn = await message.reply(
                "❌ Bạn không có quyền hủy đơn này.\n"
                "Chỉ người tạo đơn hoặc admin bot được phép hủy."
            )
            await asyncio.sleep(5)
            try:
                await bot.delete_message(message.chat.id, warn.message_id)
            except Exception:
                pass
            return

        # Lấy thông tin trước khi xóa để ghi log
        conn = db_connect()
        cur = conn.cursor()
        cur.execute("SELECT * FROM orders WHERE order_code=?", (order_code,))
        order_row = cur.fetchone()
        conn.close()

        cancel_entry = {
            "order_code": order_code,
            "cancelled_at": now_local().strftime("%Y-%m-%d %H:%M:%S"),
            "cancelled_by_id": uid,
            "cancelled_by_name": sender_display_name(message),
            "group_name": chat_title(message),
            "chat_id": message.chat.id,
            "total_amount": order_row["total_amount"] if order_row else 0,
            "total_qr": order_row["total_qr"] if order_row else 0,
        }

        # Hủy đơn: edit caption QR + xóa tin "huy"
        await cancel_order(
            bot, message.chat.id, order_code,
            cancelled_by=sender_display_name(message),
            cancel_msg_id=message.message_id,
        )

        write_cancel_log(cancel_entry)
        await notify_cancel(bot, cancel_entry)
        return

    # --- Bỏ qua command ---
    if text.startswith("/"):
        return

    # --- Kiểm tra nhóm đã kích hoạt ---
    if not chat_is_active(message.chat.id):
        logger.info("[DEBUG] Nhóm %s chưa kích hoạt — bỏ qua", message.chat.id)
        return

    # --- Chống spam: cooldown per user ---
    uid = message.from_user.id
    async with _cooldown_lock:
        last = _last_form_time[uid]
        now_ts = time.monotonic()
        if now_ts - last < FORM_COOLDOWN_SECONDS:
            return
        _last_form_time[uid] = now_ts

    # --- Parse form ---
    try:
        parsed = parse_order_form(text)
    except Exception as parse_err:
        logger.info("[DEBUG] Parse form thất bại: %s | text=%r", parse_err, text[:80])
        return

    # Kiểm tra whitelist trước khi tạo QR
    wl_errors = check_whitelist(parsed)
    if wl_errors:
        await message.reply(
            "❌ <b>Không thể tạo QR:</b>\n" + "\n".join(wl_errors),
            parse_mode=ParseMode.HTML,
        )
        return

    order_code = generate_order_code()

    try:
        sent_ids = await send_order_qrs(bot, message, order_code, parsed)
        save_order_to_db(order_code, message, parsed, sent_ids)
        logger.info("Tạo đơn thành công: %s (%d QR, %d đ)",
                    order_code, parsed["total_qr"], parsed["total_amount"])
    except Exception as e:
        logger.error("Lỗi tạo đơn %s: %s", order_code, e)
        err_msg = str(e)
        # Lỗi ngân hàng/STK → trả thẳng vào nhóm
        is_bank_err = any(k in err_msg for k in ["Ngân hàng", "STK", "không hợp lệ", "không phải ảnh", "HTTP"])
        if is_bank_err:
            await message.reply(
                f"❌ {err_msg}",
                parse_mode=ParseMode.HTML,
            )
        else:
            # Lỗi hệ thống → gửi về group noti, nhóm chỉ thấy thông báo chung
            err_text = (
                f"⚠️ <b>Lỗi tạo QR (hệ thống)</b>\n"
                f"📦 Mã đơn : <code>{order_code}</code>\n"
                f"🏘 Nhóm   : {chat_title(message)} (<code>{message.chat.id}</code>)\n"
                f"👤 User   : {sender_display_name(message)}\n"
                f"❌ Lỗi    : <pre>{err_msg}</pre>"
            )
            await notify_system(bot, err_text)
            await message.reply("❌ Có lỗi xảy ra khi tạo QR. Admin đã được thông báo.")


# ─────────────────────────── MAIN ────────────────────────────────────────────



@dp.callback_query(lambda c: c.data and c.data.startswith("sent:"))
async def cb_sent(callback: CallbackQuery, bot: Bot) -> None:
    """Xử lý khi user bấm [✅ Đã gửi đơn] — lock để tránh double-tap."""
    from aiogram.types import InlineKeyboardMarkup, InlineKeyboardButton
    order_code = callback.data.split(":", 1)[1]

    lock = await _get_order_lock(order_code)
    if lock.locked():
        await callback.answer("⏳ Đang xử lý, vui lòng đợi...", show_alert=False)
        return

    async with lock:
        conn = db_connect()
        cur = conn.cursor()
        cur.execute("SELECT status, chat_id FROM orders WHERE order_code=?", (order_code,))
        row = cur.fetchone()
        conn.close()

        if not row:
            await callback.answer("Không tìm thấy đơn.", show_alert=True)
            return
        if row["status"] != "active":
            await callback.answer("Đơn này đã được gửi hoặc xử lý rồi.", show_alert=True)
            return

    # Gửi card vào Group Collect
    collect_mid = await _send_collect_card(bot, order_code)

    # Cập nhật DB
    conn = db_connect()
    cur = conn.cursor()
    cur.execute(
        "UPDATE orders SET status='sent', collect_message_id=? WHERE order_code=?",
        (collect_mid, order_code)
    )
    conn.commit()
    conn.close()

    # Update button → [⏳ Chờ bill]
    kb = InlineKeyboardMarkup(inline_keyboard=[[
        InlineKeyboardButton(text="⏳ Chờ bill...", callback_data=f"noop:{order_code}"),
        InlineKeyboardButton(text="❌ Hủy đơn",    callback_data=f"cancel:{order_code}"),
    ]])
    try:
        await callback.message.edit_reply_markup(reply_markup=kb)
    except Exception:
        pass

        await callback.answer("✅ Đã gửi đơn sang group collect!")


@dp.callback_query(lambda c: c.data and c.data.startswith("cancel:"))
async def cb_cancel(callback: CallbackQuery, bot: Bot) -> None:
    """Xử lý khi user bấm [❌ Hủy đơn]."""
    order_code = callback.data.split(":", 1)[1]

    conn = db_connect()
    cur = conn.cursor()
    cur.execute("SELECT status FROM orders WHERE order_code=?", (order_code,))
    row = cur.fetchone()
    conn.close()

    if not row:
        await callback.answer("Không tìm thấy đơn.", show_alert=True)
        return
    if row["status"] in ("completed", "cancelled"):
        await callback.answer("Đơn này đã hoàn thành hoặc đã hủy.", show_alert=True)
        return

    cancel_entry = {}
    conn = db_connect()
    cur = conn.cursor()
    cur.execute("SELECT * FROM orders WHERE order_code=?", (order_code,))
    order_row = cur.fetchone()
    conn.close()

    if order_row:
        cancel_entry = {
            "order_code": order_code,
            "cancelled_at": now_local().strftime("%Y-%m-%d %H:%M:%S"),
            "cancelled_by_id": callback.from_user.id,
            "cancelled_by_name": callback.from_user.username or callback.from_user.full_name or str(callback.from_user.id),
            "group_name": callback.message.chat.title or str(callback.message.chat.id),
            "chat_id": callback.message.chat.id,
            "total_amount": order_row["total_amount"],
            "total_qr": order_row["total_qr"],
        }

    await cancel_order(
        bot, callback.message.chat.id, order_code,
        cancelled_by=cancel_entry.get("cancelled_by_name", "?"),
        cancel_msg_id=callback.message.message_id,
    )
    write_cancel_log(cancel_entry)
    await notify_cancel(bot, cancel_entry)
    await callback.answer("❌ Đã hủy đơn!")


@dp.callback_query(lambda c: c.data and c.data.startswith("noop:"))
async def cb_noop(callback: CallbackQuery) -> None:
    await callback.answer("Đơn đang chờ bill.", show_alert=False)


def _build_collect_html(order_code: str, status: str = "pending",
                        completer: str = "", completed_at: str = "") -> str:
    """
    Build HTML card collect — dùng nhất quán cho cả gửi lần đầu lẫn edit.
    HTML đơn giản chỉ dùng <b> và <code> để tránh lỗi entities khi edit.
    """
    conn = db_connect()
    cur = conn.cursor()
    cur.execute("SELECT * FROM orders WHERE order_code=?", (order_code,))
    o = cur.fetchone()
    cur.execute(
        "SELECT * FROM receivers WHERE order_code=? ORDER BY receiver_index",
        (order_code,)
    )
    receivers = cur.fetchall()
    conn.close()

    if not o:
        return ""

    lines = [
        f"📦 Mã đơn: <b><code>{order_code}</code></b>",
        f"💰 Tổng: <b><code>{format_money(o['total_amount'])}</code></b>  |  {o['total_qr']} QR",
        f"👤 <b>{o['sender_name']}</b>  —  <b>{(o['sender_bank'] or '').upper()}</b>  —  <code>{o['sender_account']}</code>",
        "―――――――――――――――――――――――――",
    ]
    for r in receivers:
        amt   = r["actual_amount"] if r["actual_amount"] else r["amount"]
        fixed = " ✏️" if r["actual_amount"] and r["actual_amount"] != r["amount"] else ""
        lines.append(
            f"  {r['receiver_index']}. <b>{r['receiver_name']}</b>"
            f"  |  <b>{(r['receiver_bank'] or '').upper()}</b>"
            f"  |  <code>{r['receiver_account']}</code>"
            f"  |  <b><code>{format_money(amt)}</code></b>{fixed}"
        )
    lines.append("―――――――――――――――――――――――――")
    if status == "completed":
        lines.append(f"✅ <b>Đã có bill</b>  —  {completer}  —  {completed_at}")
    else:
        lines.append("⏳ <b><u>Đang chờ bill XẢ TIỀN</u></b>, bạn iu hãy reply bill vào đây nhé 🩷")

    return "\n".join(lines)


async def _send_collect_card(bot: Bot, order_code: str) -> Optional[int]:
    """Gửi card đơn vào COLLECT_GROUP_ID dùng HTML, trả về message_id."""
    if not COLLECT_GROUP_ID:
        return None

    text = _build_collect_html(order_code)
    if not text:
        return None

    try:
        sent = await bot.send_message(
            chat_id=COLLECT_GROUP_ID,
            text=text,
            parse_mode="HTML",
        )
        return sent.message_id
    except Exception as e:
        logger.warning("Không gửi được card collect: %s", e)
        return None



# ─────────────────────────── BACKGROUND REMINDERS ────────────────────────────

async def _reminder_loop(bot: Bot) -> None:
    """
    Chạy mỗi 60 giây. Xử lý 2 loại nhắc nhở:
    A) Đơn 'active' > 5 phút chưa bấm 'Đã gửi' → tag người tạo trong Group QR
    B) Đơn 'sent' > 30 phút chưa có bill → ping Group Collect mỗi 30 phút
    """
    REMIND_SENT_AFTER   = 5 * 60        # 5 phút chưa bấm Đã gửi
    REMIND_BILL_EVERY   = 30 * 60       # nhắc mỗi 30 phút nếu chưa có bill
    reminded_sent: set[str] = set()     # order_code đã nhắc "chưa gửi"

    while True:
        await asyncio.sleep(60)
        try:
            now = now_local()
            conn = db_connect()
            cur = conn.cursor()

            # ── A: active > 5 phút chưa gửi ──────────────────────────────
            cur.execute("""
                SELECT order_code, created_at, chat_id,
                       creator_username, button_message_id
                FROM orders
                WHERE status = 'active'
                  AND date(created_at) = date('now', 'localtime')
            """)
            active_rows = cur.fetchall()
            conn.close()

            for row in active_rows:
                oc = row["order_code"]
                if oc in reminded_sent:
                    continue
                try:
                    created = datetime.strptime(row["created_at"], "%Y-%m-%d %H:%M:%S")
                    elapsed = (now - created).total_seconds()
                    if elapsed >= REMIND_SENT_AFTER:
                        tag = row["creator_username"] or "bạn iu"
                        try:
                            await bot.send_message(
                                chat_id=row["chat_id"],
                                text=(
                                    f"⏰ {tag} ơi, đơn <code>{oc}</code> "
                                    f"đã được tạo {int(elapsed//60)} phút rồi mà chưa bấm "
                                    f"<b>Đã gửi đơn</b> nha! 🥺"
                                ),
                                parse_mode="HTML",
                                reply_to_message_id=row["button_message_id"],
                            )
                            reminded_sent.add(oc)
                        except Exception as e:
                            logger.warning("Không nhắc được đơn %s: %s", oc, e)
                except Exception:
                    pass

            # ── B: sent > 30 phút chưa có bill ───────────────────────────
            conn = db_connect()
            cur = conn.cursor()
            cur.execute("""
                SELECT order_code, created_at, collect_message_id,
                       total_amount, total_qr
                FROM orders
                WHERE status = 'sent'
                  AND collect_message_id IS NOT NULL
                  AND date(created_at) = date('now', 'localtime')
            """)
            sent_rows = cur.fetchall()
            conn.close()

            for row in sent_rows:
                oc = row["order_code"]
                try:
                    created = datetime.strptime(row["created_at"], "%Y-%m-%d %H:%M:%S")
                    elapsed = (now - created).total_seconds()
                    # Nhắc mỗi 30 phút (30p, 60p, 90p...)
                    intervals = int(elapsed // REMIND_BILL_EVERY)
                    if intervals >= 1:
                        # Tránh gửi trùng trong cùng 1 phút
                        remainder = elapsed % REMIND_BILL_EVERY
                        if remainder <= 60:   # trong vòng 60s đầu của mỗi interval
                            try:
                                await bot.send_message(
                                    chat_id=COLLECT_GROUP_ID,
                                    text=(
                                        f"⚠️ Đơn <code>{oc}</code> đã gửi "
                                        f"<b>{int(elapsed//60)} phút</b> rồi mà chưa có bill nha! 🩷\n"
                                        f"💰 {format_money(row['total_amount'])}  |  {row['total_qr']} QR"
                                    ),
                                    parse_mode="HTML",
                                    reply_to_message_id=row["collect_message_id"],
                                )
                            except Exception as e:
                                logger.warning("Không nhắc bill đơn %s: %s", oc, e)
                except Exception:
                    pass

            # Cleanup _callback_locks cũ (giữ memory)
            old_codes = set()
            conn_cl = db_connect()
            cur_cl = conn_cl.cursor()
            cur_cl.execute(
                "SELECT order_code FROM orders WHERE status IN ('completed','cancelled','expired')"
            )
            old_codes = {r["order_code"] for r in cur_cl.fetchall()}
            conn_cl.close()
            async with _callback_locks_mu:
                for oc in list(_callback_locks.keys()):
                    if oc in old_codes:
                        del _callback_locks[oc]

        except Exception as e:
            logger.error("_reminder_loop lỗi: %s", e)





async def _daily_report_loop(bot: Bot) -> None:
    """
    Tự động xuất báo cáo Excel cuối ngày lúc 23:30 GMT+7.
    Mỗi Group QR nhận file riêng chỉ chứa đơn của group đó.
    Đồng thời gửi bản tổng vào COLLECT_GROUP_ID.
    """
    REPORT_HOUR   = 0
    REPORT_MINUTE = 0
    last_report_date: Optional[str] = None

    while True:
        try:
            now = now_local()
            today_str = now.strftime("%d/%m/%Y")

            if (now.hour == REPORT_HOUR and now.minute >= REPORT_MINUTE
                    and last_report_date != today_str):

                # Báo cáo lúc 00:00 → lấy dữ liệu ngày HÔM QUA
                from datetime import timedelta
                report_date = (now - timedelta(days=1)).strftime("%d/%m/%Y")
                safe_date   = (now - timedelta(days=1)).strftime("%d-%m-%Y")
                cur = conn.cursor()
                cur.execute("SELECT DISTINCT chat_id FROM activated_chats")
                active_chats = [r["chat_id"] for r in cur.fetchall()]
                conn.close()

                has_any = False

                for chat_id in active_chats:
                    # Lấy đơn của riêng group này
                    rows = fetch_by_date(report_date, chat_id=chat_id)
                    if not rows:
                        continue

                    has_any = True
                    total_orders = len(set(r["order_code"] for r in rows))
                    total_amount = sum(r["amount"] for r in rows)
                    completed    = sum(1 for r in rows if r["status"] == "completed")
                    cancelled    = sum(1 for r in rows if r["status"] == "cancelled")
                    group_name   = rows[0]["group_name"] or str(chat_id)

                    caption = (
                        f"📊 <b>Báo cáo ngày {report_date}</b>\n"
                        f"🏘 {group_name}\n"
                        f"━━━━━━━━━━━━━━━━━━━━\n"
                        f"📦 Tổng đơn   : <b>{total_orders}</b>\n"
                        f"✅ Hoàn thành : <b>{completed}</b>\n"
                        f"❌ Đã hủy     : <b>{cancelled}</b>\n"
                        f"⏳ Chưa xử lý : <b>{total_orders - completed - cancelled}</b>\n"
                        f"💰 Tổng tiền  : <b><code>{format_money(total_amount)}</code></b>"
                    )

                    with tempfile.TemporaryDirectory() as tmpdir:
                        fname_xlsx = f"baocao_{safe_date}.xlsx"
                        fpath = os.path.join(tmpdir, fname_xlsx)
                        export_orders_to_excel(rows, fpath)

                        from aiogram.types import FSInputFile

                        # Gửi về đúng Group QR
                        try:
                            await bot.send_document(
                                chat_id=chat_id,
                                document=FSInputFile(fpath, filename=fname_xlsx),
                                caption=caption,
                                parse_mode="HTML",
                            )
                        except Exception as e:
                            logger.warning("Gửi báo cáo group %s thất bại: %s", chat_id, e)

                    logger.info("Báo cáo %s group %s: %d đơn", today_str, chat_id, total_orders)

                if not has_any:
                    await notify_system(bot,
                        f"📊 <b>Báo cáo ngày {report_date}</b>\n\nKhông có đơn nào trong ngày hôm nay.")

                last_report_date = today_str

        except Exception as e:
            logger.error("_daily_report_loop lỗi: %s", e)

        await asyncio.sleep(60)


async def main() -> None:
    init_db()
    bot = Bot(
        token=BOT_TOKEN,
        default=DefaultBotProperties(parse_mode=ParseMode.HTML),
    )
    start_time = now_local().strftime("%Y-%m-%d %H:%M:%S")
    logger.info("QrTC V2 khởi động — superadmins: %s", SUPER_ADMIN_IDS)
    # Gửi thông báo khởi động đến group hệ thống
    async def _send_startup():
        await asyncio.sleep(2)
        # Auto-expire đơn cũ từ ngày trước khi khởi động
        conn = db_connect()
        cur = conn.cursor()
        cur.execute("""
            UPDATE orders SET status = 'expired'
            WHERE status IN ('active', 'sent')
              AND date(created_at) < date('now', 'localtime')
        """)
        expired_count = cur.rowcount
        conn.commit()
        conn.close()
        if expired_count:
            logger.info("Auto-expired %d đơn cũ khi khởi động", expired_count)
        await notify_system(
            bot,
            f"🟢 <b>QrTC V2 đã khởi động</b>\n"
            f"🕐 Lúc : {start_time}\n"
            f"👑 Superadmin : {', '.join(str(x) for x in sorted(SUPER_ADMIN_IDS))}"
        )
    asyncio.create_task(_send_startup())
    asyncio.create_task(_reminder_loop(bot))        # nhắc 5p + 30p
    asyncio.create_task(_daily_report_loop(bot))   # báo cáo 23:30 GMT+7
    await dp.start_polling(bot)


if __name__ == "__main__":
    asyncio.run(main())
